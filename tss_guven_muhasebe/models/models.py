
# -*- coding: utf-8 -*-
from odoo import models, fields, api, tools, _
from odoo.exceptions import UserError, ValidationError
from datetime import datetime, timedelta
import requests
from zeep import Client, Settings
from zeep.transports import Transport
import xml.etree.ElementTree as ET
import logging

_logger = logging.getLogger(__name__)

try:
    import pymssql
except ImportError:
    _logger.warning("pymssql kütüphanesi bulunamadı. Logo senkronizasyonu için 'pip install pymssql' komutunu çalıştırın.")
    pymssql = None


class e_invoice(models.Model):
    _name = 'e.invoice'
    _description = 'E-Fatura Kayıtları'
    _order = 'issue_date, invoice_id'
    _rec_name = 'invoice_id'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    # SQL Constraints (v1.0.6)
    _sql_constraints = [
        ('unique_invoice_record',
         'UNIQUE(invoice_id, kaynak, uuid)',
         'Bu fatura kaydı zaten mevcut! (Aynı invoice_id, kaynak ve UUID kombinasyonu tekrar edilemez.)')
    ]

    # Ana Bilgiler
    invoice_id = fields.Char(string='Fatura ID', required=True, index=True)
    uuid = fields.Char(string='UUID', required=True, index=True)
    
    # Header Bilgileri
    sender = fields.Char(string='Gönderen')
    receiver = fields.Char(string='Alıcı')
    supplier = fields.Char(string='Tedarikçi')
    customer = fields.Char(string='Müşteri')
    
    # Tarih ve Durum
    issue_date = fields.Datetime(string='Fatura Tarihi')
    create_date_ws = fields.Datetime(string='Oluşturma Tarihi (WS)')
    
    # Finansal Bilgiler
    payable_amount = fields.Float(string='Ödenecek Tutar', digits=(16, 2))
    tax_exclusive_total_amount = fields.Float(string='Vergi Hariç Toplam', digits=(16, 2))
    tax_inclusive_total_amount = fields.Float(string='Vergi Dahil Toplam', digits=(16, 2))
    allowance_total_amount = fields.Float(string='İndirim Tutarı', digits=(16, 2))
    line_extension_amount = fields.Float(string='Satır Uzantı Tutarı', digits=(16, 2))
    
    profile_id = fields.Char(string='Profil ID')
    invoice_type_code = fields.Char(string='Fatura Tip Kodu')
    status = fields.Char(string='Durum', tracking=True, help="E-fatura durumu")
    status_description = fields.Char(string='Durum Açıklaması', tracking=True)
    status_code = fields.Char(string='Durum Kodu')
    status_detail = fields.Char(string='Durum Detayı', help="Detaylı durum açıklaması", compute='_get_status_detail', store=True)
    gib_status_code = fields.Char(string='GİB Durum Kodu', tracking=True)
    gib_status_description = fields.Char(string='GİB Durum Açıklaması')

    # Yanıt Bilgileri (v1.0.4)
    response_code = fields.Char(
        string='Yanıt Kodu',
        tracking=True,
        help="Fatura yanıt kodu (kabul/red/iptal bilgisi)"
    )
    response_description = fields.Char(
        string='Yanıt Açıklaması',
        tracking=True,
        help="Fatura yanıt açıklaması (detaylı durum)"
    )

    envelope_identifier = fields.Char(string='Zarf Tanımlayıcı')
    direction = fields.Selection([
        ('IN', 'Gelen'),
        ('OUT', 'Giden')
    ], string='Yön', default='IN', tracking=True)
    
    # İlişkiler
    from_field = fields.Char(string='Kimden')
    to_field = fields.Char(string='Kime')
    
    # Aktif/Pasif
    active = fields.Boolean(string='Arşivlenmedi', default=True, tracking=True)
    gvn_active = fields.Boolean(string='Geçerli Faturalar', compute='_compute_active', store=True)
    exists_in_logo = fields.Boolean(string='Logo\'da Var', default=False, help="Logo entegrasyonunda bu fatura var mı?")
    logo_record_id = fields.Integer(string='Logo Kayıt ID', help="Logo entegrasyonunda bu faturanın ID'si")

    # Kaynak Bilgisi
    kaynak = fields.Selection([
        ('e-fatura', 'E-Fatura'),
        ('e-arsiv', 'E-Arşiv'),
        ('e-irsaliye', 'E-İrsaliye'),
    ], string='Kaynak', default='e-fatura', required=True, index=True, tracking=True, help="Belgenin kaynağı")

    # E-Arşiv Özel Alanları
    currency_code = fields.Char(
        string='Para Birimi',
        size=3,
        help="Para birimi kodu (TRY, USD, EUR, etc.)"
    )
    reported = fields.Boolean(
        string='Raporlandı',
        default=False,
        help="E-Arşiv faturası raporlandı mı?"
    )
    earchive_type = fields.Char(
        string='E-Arşiv Tipi',
        help="E-Arşiv fatura tipi (EARCHIVE_TYPE)"
    )
    sending_type = fields.Char(
        string='Gönderim Tipi',
        help="E-Arşiv gönderim tipi (SENDING_TYPE)"
    )

    # Notlar
    notes = fields.Text(string='Notlar')

    # Kilit Mekanizması (v1.0.5)
    is_locked = fields.Boolean(
        string='Kilitli',
        default=False,
        tracking=True,
        help="Bu kayıt kilitlendiğinde otomatik senkronizasyonlar ve manuel düzenlemeler engellenecektir."
    )
    locked_by_id = fields.Many2one(
        'res.users',
        string='Kilitleyen Kullanıcı',
        readonly=True,
        tracking=True,
        help="Kaydı kilitleyen kullanıcı"
    )
    locked_date = fields.Datetime(
        string='Kilitlenme Tarihi',
        readonly=True,
        tracking=True,
        help="Kaydın kilitlendiği tarih ve saat"
    )
    lock_reason = fields.Text(
        string='Kilitleme Nedeni',
        help="Kaydın neden kilitlendiği (opsiyonel)"
    )

    @api.depends('status_code', 'kaynak', 'profile_id')
    def _compute_active(self):
        """Geçerli faturaları belirle - kaynak tipine göre farklı mantık (v1.0.6)"""
        for record in self:
            if record.kaynak == 'e-arsiv':
                # ÖNCE profile_id kontrolü (daha kesin ve öncelikli)
                if record.profile_id == 'IPTAL':
                    record.gvn_active = False
                # Sonra status_code kontrolü
                # 150: RAPORLANMADAN İPTAL EDİLDİ
                # 200: FATURA ID BULUNAMADI
                elif record.status_code in ['150', '200']:
                    record.gvn_active = False
                else:
                    record.gvn_active = True
            else:
                # E-Fatura için geçersiz status kodları
                # 116: Geçersiz, 120: Ret Edildi, 130: Reddedildi, 136: GİB'e Gönderim Hatası
                record.gvn_active = record.status_code not in ['116', '120', '130', '136']

    @api.depends('status_code', 'kaynak')
    def _get_status_detail(self):
        """Durum kodunu açıklamaya çevir - kaynak tipine göre farklı açıklamalar"""
        # E-Fatura durum kodları
        efatura_code_details = {
            # Giden Fatura Durumları
            '100': 'Durum Sorgulanması Yapmaya Devam Edilecek',
            '101': 'Fatura Yükleme - Başarılı',
            '102': 'Belge İşleniyor',
            '103': 'Belge GİB\'e Göndermek İçin Zarflanıyor',
            '104': 'Belge Zarflandı GİB\'e Gönderilecek',
            '105': 'Belge Zarflanırken Hata Oluştu. Tekrar Denenecektir.',
            '106': 'Belge İmzalanıyor',
            '107': 'Belge İmzalandı',
            '109': 'Belge GİB\'e Gönderildi',
            '110': 'Belge Alıcıya Başarıyla Ulaştırıldı. Sistem Yanıtı Bekliyor.',
            '111': 'Ticari Belge Alıcıdan Onay Bekliyor',
            '112': 'Belge Kabul Edildi',
            '116': 'izibiz Referans Kodu Değil (Muhtemel Geçersiz - 116)',
            '117': 'Red Alıcıdan Yanıt Bekliyor',
            '120': 'Belge Ret Edildi',
            '134': 'Belge GİB\'e Gönderilirken Zaman Aşımına Uğradı.',
            '135': 'Belge GİB\'e Gönderiliyor',
            '136': 'Belge GİB\'e Gönderilirken Hata Oluştu',
            '137': 'Belge GİB\'e Gönderildi',
            '139': 'Otomatik Gönderim Hatası',
            '140': 'Belge Numarası Atandı',
            '141': 'Belge Numarası Atandı',

            # Gelen Fatura Durumları
            '133': 'Temel Fatura Alındı',
            '132': 'Ticari Fatura Yanıt Bekliyor',
            '122': 'Kabul Edildi',
            '123': 'Kabul İşleniyor',
            '124': 'Kabul GİB\'den Yanıt Bekliyor',
            '125': 'Kabul Alıcıdan Yanıt Bekliyor',
            '126': 'Kabul İşlemi Başarısız',
            '127': 'Red Alıcıdan Yanıt Bekliyor',
            '128': 'Red GİB\'de Yanıt Bekliyor',
            '129': 'Red İşleniyor',
            '130': 'Reddedildi',
            '131': 'Red İşlemi Başarısız',
        }

        # E-Arşiv durum kodları
        earsiv_code_details = {
            '100': 'Kuyruğa Eklendi',
            '105': 'Taslak Olarak Eklendi',
            '110': 'İşleniyor',
            '120': 'Raporlanacak',
            '130': 'Raporlandı',  # ✅ E-Arşiv için başarılı durum
            '150': 'Raporlanmadan İptal Edildi. Raporlanmayacak.',
            '200': 'Fatura ID Bulunamadı',
        }

        for record in self:
            if record.kaynak == 'e-arsiv':
                # E-Arşiv için özel durum kodları
                if record.status_code in earsiv_code_details:
                    record.status_detail = earsiv_code_details[record.status_code]
                else:
                    record.status_detail = f'Bilinmeyen Durum ({record.status_code})'
            else:
                # E-Fatura için durum kodları
                if record.status_code in efatura_code_details:
                    record.status_detail = efatura_code_details[record.status_code]
                else:
                    record.status_detail = f'Bilinmeyen Durum ({record.status_code})'

    @api.model
    def _parse_date_field(self, date_string, field_name="tarih"):
        """
        Çeşitli tarih formatlarını parse etmek için yardımcı metod
        
        Args:
            date_string (str): Parse edilecek tarih string'i
            field_name (str): Hata logları için alan adı
            
        Returns:
            datetime: Parse edilmiş naive datetime objesi veya None
        """
        if not date_string:
            return None
        
        # String'i temizle
        date_string = str(date_string).strip()
            
        try:
            # ÖZEL DURUM: ISSUE_DATE için sadece tarih + timezone formatı (YYYY-MM-DD+HH:MM)
            # Örnek: 2025-01-01+03:00 → Sadece tarih kısmını al (saat bilgisi yok)
            if '+' in date_string and 'T' not in date_string:
                date_part = date_string.split('+')[0]
                if len(date_part) == 10:  # YYYY-MM-DD (sadece tarih)
                    # Timezone kısmını ignore et, sadece tarihi al (00:00:00 olarak)
                    try:
                        return datetime.fromisoformat(date_part)
                    except ValueError:
                        pass
                # Eğer date_part 10 karakterden fazlaysa (saat bilgisi varsa)
                # O zaman eski yöntemi kullan (aşağıya düşecek)
            
            # ISO format kontrolü (öncelikli - SOAP servislerde yaygın)
            if 'T' in date_string:
                # Timezone bilgisini temizle veya çevir
                if date_string.endswith('Z'):
                    clean_date = date_string.replace('Z', '+00:00')
                else:
                    clean_date = date_string
                
                parsed_dt = datetime.fromisoformat(clean_date)
                
                # Eğer timezone aware ise UTC'ye çevir ve naive yap
                if parsed_dt.tzinfo is not None:
                    utc_dt = parsed_dt.utctimetuple()
                    return datetime(*utc_dt[:6])
                else:
                    return parsed_dt
            
            # Timezone offset'li TAM tarih-saat kontrolü (YYYY-MM-DDTHH:MM:SS+HH:MM formatı)
            if 'T' in date_string and ('+' in date_string or date_string.endswith('Z')):
                try:
                    parsed_dt = datetime.fromisoformat(date_string)
                    # UTC'ye çevir ve naive yap
                    if parsed_dt.tzinfo is not None:
                        utc_dt = parsed_dt.utctimetuple()
                        return datetime(*utc_dt[:6])
                    else:
                        return parsed_dt
                except ValueError:
                    # Z suffix varsa +00:00 ile değiştir
                    if date_string.endswith('Z'):
                        parsed_dt = datetime.fromisoformat(date_string.replace('Z', '+00:00'))
                        utc_dt = parsed_dt.utctimetuple()
                        return datetime(*utc_dt[:6])
            
            # Sadece tarih kısmı var ise (YYYY-MM-DD)
            if len(date_string) == 10 and '-' in date_string:
                return datetime.fromisoformat(date_string)
                
            # Diğer yaygın formatları dene
            date_formats = [
                '%d-%m-%Y %H:%M:%S',  # 02-01-2025 15:23:28 (E-Arşiv format)
                '%Y-%m-%d %H:%M:%S',  # 2025-05-02 14:30:00
                '%Y-%m-%dT%H:%M:%S',  # 2025-05-02T14:30:00
                '%Y-%m-%d',           # 2025-05-02
                '%d.%m.%Y',           # 02.05.2025
                '%d/%m/%Y',           # 02/05/2025
                '%d-%m-%Y',           # 02-05-2025
                '%Y%m%d',             # 20250502
            ]
            
            for date_format in date_formats:
                try:
                    return datetime.strptime(date_string, date_format)
                except ValueError:
                    continue
            
            # Hiçbiri işe yaramadıysa hata logla
            _logger.error("E-Fatura: %s alanı parse edilemedi: %s", field_name, date_string)
            return None
            
        except Exception as e:
            _logger.error("E-Fatura: %s parse hatası: %s - %s", field_name, date_string, str(e))
            return None

    @api.model
    def _parse_financial_field(self, value_string, field_name="tutar"):
        """
        Finansal alanları güvenli şekilde float'a dönüştür
        
        Args:
            value_string (str): Parse edilecek değer
            field_name (str): Hata logları için alan adı
            
        Returns:
            float: Parse edilmiş değer veya 0.0
        """
        if not value_string:
            return 0.0
            
        try:
            # String temizleme
            clean_value = str(value_string).strip()
            
            # Türkçe ondalık ayırıcı (virgül) kontrolü
            if ',' in clean_value and '.' in clean_value:
                # Hem virgül hem nokta var - 1.234,56 formatı
                clean_value = clean_value.replace('.', '').replace(',', '.')
            elif ',' in clean_value:
                # Sadece virgül var - 1234,56 formatı
                clean_value = clean_value.replace(',', '.')
            
            # Para birimi sembollerini temizle
            clean_value = clean_value.replace('₺', '').replace('TL', '').replace('$', '').strip()
            
            # Float'a dönüştür
            return float(clean_value)
            
        except (ValueError, TypeError) as e:
            _logger.warning("E-Fatura: %s alanı parse edilemedi: %s - Hata: %s", field_name, value_string, str(e))
            return 0.0

    def test_date_parsing(self):
        """Tarih parse işlemlerini test et - geliştirme amaçlı"""
        test_dates = [
            '2025-05-02+03:00',      # SOAP servisinizden gelen format
            '2025-05-02T14:30:00Z',  # ISO 8601 with Z
            '2025-05-02T14:30:00+03:00',  # ISO 8601 with timezone
            '2025-05-02',            # Sadece tarih
            '02.05.2025',            # Türkçe format
            '02/05/2025',            # Alternatif
        ]
        
        for test_date in test_dates:
            parsed = self._parse_date_field(test_date, 'test_{}'.format(test_date))
            _logger.info("Test: %s -> %s", test_date, parsed)

        return True

    # Kilit Mekanizması Metodları (v1.0.5)
    def write(self, vals):
        """Kilitli kayıtları korumak için write metodunu override et"""
        # Kilit alanlarının kendisi güncelleniyorsa izin ver
        lock_fields = {'is_locked', 'locked_by_id', 'locked_date', 'lock_reason'}
        if set(vals.keys()).issubset(lock_fields):
            return super(e_invoice, self).write(vals)

        # Kilitli kayıtları kontrol et
        locked_records = self.filtered('is_locked')
        if locked_records:
            raise UserError(
                f"{len(locked_records)} adet kilitli kayıt var!\n\n"
                f"Kilitli fatura ID'leri: {', '.join(locked_records.mapped('invoice_id'))}\n\n"
                "Lütfen önce kayıtların kilidini kaldırın."
            )

        return super(e_invoice, self).write(vals)

    def action_lock(self):
        """Kayıtları kilitle"""
        for record in self:
            if not record.is_locked:
                record.write({
                    'is_locked': True,
                    'locked_by_id': self.env.user.id,
                    'locked_date': fields.Datetime.now(),
                })
                # Chatter'a mesaj ekle
                record.message_post(
                    body=f"Kayıt {self.env.user.name} tarafından kilitlendi.",
                    subject="Kayıt Kilitlendi"
                )

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Başarılı',
                'message': f'{len(self)} kayıt kilitlendi.',
                'type': 'success',
                'sticky': False,
            }
        }

    def action_unlock(self):
        """Kayıtların kilidini kaldır"""
        for record in self:
            if record.is_locked:
                locked_by = record.locked_by_id.name if record.locked_by_id else 'Bilinmeyen'
                locked_date = record.locked_date.strftime('%d.%m.%Y %H:%M') if record.locked_date else 'Bilinmeyen'

                record.write({
                    'is_locked': False,
                    'locked_by_id': False,
                    'locked_date': False,
                    'lock_reason': False,
                })
                # Chatter'a mesaj ekle
                record.message_post(
                    body=f"Kilidin açılması: {self.env.user.name}<br/>"
                         f"Daha önce kilitleyen: {locked_by} ({locked_date})",
                    subject="Kilit Kaldırıldı"
                )

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Başarılı',
                'message': f'{len(self)} kayıdın kilidi kaldırıldı.',
                'type': 'success',
                'sticky': False,
            }
        }

    @api.model
    def create_from_soap_data(self, soap_data):
        """SOAP servisinden gelen veriyi Odoo modeline dönüştür"""
        invoice_vals = {}

        if isinstance(soap_data, dict):
            # Ana bilgiler
            invoice_vals['invoice_id'] = soap_data.get('ID')
            invoice_vals['uuid'] = soap_data.get('UUID')
            invoice_vals['direction'] = soap_data.get('direction', 'IN')
            invoice_vals['kaynak'] = 'e-fatura'  # E-Fatura kaynağı
            
            # Header bilgilerini işle
            header = soap_data.get('HEADER', {})
            if header:
                invoice_vals.update({
                    'sender': header.get('SENDER'),
                    'receiver': header.get('RECEIVER'),
                    'supplier': header.get('SUPPLIER'),
                    'customer': header.get('CUSTOMER'),
                    'profile_id': header.get('PROFILEID'),
                    'invoice_type_code': header.get('INVOICE_TYPE_CODE'),
                    'status': header.get('STATUS'),
                    'status_description': header.get('STATUS_DESCRIPTION'),
                    'status_code': header.get('STATUS_CODE'),
                    'gib_status_code': header.get('GIB_STATUS_CODE'),
                    'gib_status_description': header.get('GIB_STATUS_DESCRIPTION'),
                    'response_code': header.get('RESPONSE_CODE'),  # YENİ (v1.0.4)
                    'response_description': header.get('RESPONSE_DESCRIPTION'),  # YENİ (v1.0.4)
                    'envelope_identifier': header.get('ENVELOPE_IDENTIFIER'),
                    'from_field': header.get('FROM'),
                    'to_field': header.get('TO'),
                })
                
                # Tarihleri dönüştür - Yardımcı metod kullan
                if header.get('ISSUE_DATE'):
                    parsed_date = self._parse_date_field(header.get('ISSUE_DATE'), 'ISSUE_DATE')
                    if parsed_date:
                        invoice_vals['issue_date'] = parsed_date
                
                if header.get('CDATE'):
                    parsed_date = self._parse_date_field(header.get('CDATE'), 'CDATE')
                    if parsed_date:
                        invoice_vals['create_date_ws'] = parsed_date
                
                # Finansal tutarları dönüştür - Yardımcı metod kullan
                financial_field_mapping = {
                    'PAYABLE_AMOUNT': 'payable_amount',
                    'TAX_EXCLUSIVE_TOTAL_AMOUNT': 'tax_exclusive_total_amount',
                    'TAX_INCLUSIVE_TOTAL_AMOUNT': 'tax_inclusive_total_amount',
                    'ALLOWANCE_TOTAL_AMOUNT': 'allowance_total_amount',
                    'LINE_EXTENSION_AMOUNT': 'line_extension_amount'
                }
                
                for soap_field, odoo_field in financial_field_mapping.items():
                    if header.get(soap_field):
                        parsed_amount = self._parse_financial_field(header.get(soap_field), soap_field)
                        invoice_vals[odoo_field] = parsed_amount
        
        return self.create(invoice_vals)

    @api.model
    def _get_soap_client_and_login(self):
        """
        Ortak SOAP client oluşturma ve login metodu
        E-Fatura ve E-Arşiv için ortak kullanılabilir

        Returns:
            tuple: (efatura_client, session_id, transport, settings)
        """
        soap_config = self.env['ir.config_parameter'].sudo()
        username = soap_config.get_param('efatura.username')
        password = soap_config.get_param('efatura.password')
        efatura_ws = soap_config.get_param('efatura.ws_url',
            'https://efaturaws.izibiz.com.tr/EInvoiceWS?wsdl')

        if not username or not password:
            raise UserError(_("SOAP kimlik bilgileri yapılandırılmamış!"))

        # SOAP transport ayarları
        session_ws = requests.Session()
        session_ws.verify = True
        transport_ws = Transport(session=session_ws, timeout=60)
        settings_ws = Settings(strict=False, xml_huge_tree=True, forbid_dtd=False, forbid_entities=False)

        # E-Fatura Client oluştur
        efatura_client = Client(wsdl=efatura_ws, transport=transport_ws, settings=settings_ws)

        # Login
        request_header = {
            'SESSION_ID': '-1',
            'APPLICATION_NAME': 'Odoo SOAP Client',
        }

        login_response = efatura_client.service.Login(
            REQUEST_HEADER=request_header,
            USER_NAME=username,
            PASSWORD=password
        )
        session_id = login_response.SESSION_ID
        _logger.info("SOAP Login başarılı. Session ID: %s", session_id)

        return efatura_client, session_id, transport_ws, settings_ws

    @api.model
    def sync_invoices_from_soap(self, start_date, end_date, direction='IN'):
        """SOAP servisinden faturaları senkronize et - Logo otomatik sync ile"""
        try:
            # Ortak SOAP client ve login
            efatura_client, session_id, transport_ws, settings_ws = self._get_soap_client_and_login()

            # Fatura listesini al
            request_header = {'SESSION_ID': session_id}

            # Tarihleri datetime objesine çevir (SOAP servisi datetime bekliyor)
            from datetime import datetime as dt
            try:
                start_dt = dt.strptime(start_date, '%Y-%m-%d')
                end_dt = dt.strptime(end_date, '%Y-%m-%d')
            except:
                # Eğer parse edemezse bugünden başla
                end_dt = dt.now()
                start_dt = end_dt - timedelta(days=7)

            search_key = {
                'LIMIT': '25000',
                'START_DATE': start_dt,
                'END_DATE': end_dt,
                'READ_INCLUDED': 'true',
                'DIRECTION': direction,
            }
            
            with efatura_client.settings(raw_response=True):
                raw_xml_response = efatura_client.service.GetInvoice(
                    REQUEST_HEADER=request_header,
                    INVOICE_SEARCH_KEY=search_key,
                    HEADER_ONLY='Y',
                )
                
                # XML'i parse et
                root = ET.fromstring(raw_xml_response.content)
                invoices = []

                for invoice_elem in root.findall('.//INVOICE') or root.findall('.//*[local-name()="INVOICE"]'):
                    invoice_data = {}
                    header_elem = invoice_elem.find('HEADER') or invoice_elem.find('.//*[local-name()="HEADER"]')

                    if header_elem is not None:
                        header_data = {}
                        for child in header_elem:
                            tag_name = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                            header_data[tag_name] = child.text
                        invoice_data['HEADER'] = header_data

                    invoice_data['ID'] = invoice_elem.get('ID')
                    invoice_data['UUID'] = invoice_elem.get('UUID')
                    invoice_data['direction'] = direction
                    invoices.append(invoice_data)
            
            # Logout
            efatura_client.service.Logout(REQUEST_HEADER=request_header)
            
            # Odoo'ya kaydet
            created_count = 0
            updated_count = 0
            
            for invoice_data in invoices:
                # UUID unique olduğu için tek başına yeterli (v1.0.6 performans optimizasyonu)
                existing_invoice = self.search([
                    ('uuid', '=', invoice_data.get('UUID'))
                ], limit=1)

                if existing_invoice:
                    # Kilitli kayıtları güncelleme (v1.0.5)
                    if existing_invoice.is_locked:
                        _logger.info(f"SOAP Sync: Kilitli kayıt atlandı - {existing_invoice.invoice_id}")
                        continue

                    # Güncelle
                    invoice_vals = self._prepare_invoice_vals_from_soap(invoice_data)
                    existing_invoice.write(invoice_vals)
                    updated_count += 1
                else:
                    # Yeni oluştur
                    self.create_from_soap_data(invoice_data)
                    created_count += 1
            
            result = {
                'success': True,
                'created': created_count,
                'updated': updated_count,
                'message': _('%s yeni fatura oluşturuldu, %s fatura güncellendi.') % (created_count, updated_count)
            }
            
            # Otomatik Logo senkronizasyonu kontrolü
            config_param = self.env['ir.config_parameter'].sudo()
            logo_auto_sync = config_param.get_param('logo.auto_sync', False)

            if logo_auto_sync and result.get('success'):
                try:
                    if not pymssql:
                        _logger.warning("Otomatik Logo Sync: pymssql yüklü değil")
                        result['message'] += _('\n\nLogo Senkronizasyonu: pymssql eksik')
                    else:
                        # MSSQL config al
                        server = config_param.get_param('logo.mssql_server')
                        port = int(config_param.get_param('logo.mssql_port', '1433'))
                        database = config_param.get_param('logo.mssql_database')
                        username = config_param.get_param('logo.mssql_username')
                        password = config_param.get_param('logo.mssql_password')
                        table_name = config_param.get_param('logo.invoice_table', 'LG_600_01_INVOICE')

                        if not all([server, database, username, password]):
                            _logger.warning("Otomatik Logo Sync: MSSQL config eksik")
                            result['message'] += _('\n\nLogo Senkronizasyonu: MSSQL config eksik')
                        else:
                            # Filtrelenmiş faturaları al
                            domain = [
                                ('issue_date', '>=', start_date),
                                ('issue_date', '<=', end_date),
                                ('direction', '=', direction),
                                ('kaynak', 'in', ['e-fatura', 'e-arsiv'])
                            ]
                            invoices = self.search(domain)

                            if invoices:
                                # MSSQL bağlantısı
                                conn = pymssql.connect(
                                    server=server, port=port, user=username,
                                    password=password, database=database,
                                    timeout=30, login_timeout=30, charset='UTF-8'
                                )
                                cursor = conn.cursor()
                                stats = {'found': 0, 'not_found': 0, 'errors': 0}

                                # TRCODE belirle
                                if direction == 'IN':
                                    trcode_condition = "TRCODE IN (1,3,4,13)"
                                else:  # OUT
                                    trcode_condition = "TRCODE IN (6,7,8,9,14)"

                                # Her faturayı kontrol et
                                for invoice in invoices:
                                    try:
                                        # Tarih kontrolü için issue_date'i al
                                        invoice_date = invoice.issue_date.strftime('%Y-%m-%d') if invoice.issue_date else None

                                        if not invoice_date:
                                            _logger.warning("Logo sync fatura %s: Tarih yok, atlanıyor", invoice.invoice_id)
                                            continue

                                        # Standart sorgu
                                        query = """
                                            SELECT LOGICALREF
                                            FROM {}
                                            WHERE (FICHENO = %s OR DOCODE = %s)
                                            AND CANCELLED = 0
                                            AND CAST(DATE_ AS DATE) = %s
                                            AND {}
                                        """.format(table_name, trcode_condition)

                                        cursor.execute(query, (invoice.invoice_id, invoice.invoice_id, invoice_date))
                                        logo_result = cursor.fetchone()

                                        if logo_result:
                                            invoice.write({
                                                'exists_in_logo': True,
                                                'logo_record_id': logo_result[0]
                                            })
                                            stats['found'] += 1
                                        else:
                                            invoice.write({'exists_in_logo': False})
                                            stats['not_found'] += 1
                                    except Exception as e:
                                        stats['errors'] += 1
                                        _logger.error("Logo sync fatura %s: %s", invoice.invoice_id, str(e))

                                conn.close()
                                result['message'] += _('\n\nLogo Sync: %d bulundu, %d bulunamadı') % (stats['found'], stats['not_found'])
                            else:
                                result['message'] += _('\n\nLogo Sync: Fatura bulunamadı')

                except Exception as e:
                    import traceback
                    error_trace = traceback.format_exc()
                    _logger.warning("Otomatik Logo senkronizasyonu başarısız: %s\n%s", str(e), error_trace)
                    result['message'] += _('\n\nLogo Senkronizasyonu: Otomatik sync başarısız - %s') % str(e)
            
            return result
            
        except Exception as e:
            _logger.error("SOAP senkronizasyon hatası: %s", str(e))
            return {
                'success': False,
                'error': str(e)
            }

    def _prepare_invoice_vals_from_soap(self, soap_data):
        """SOAP verisini Odoo vals formatına dönüştür (UPDATE için)"""
        vals = {}
        header = soap_data.get('HEADER', {})

        if header:
            vals.update({
                'sender': header.get('SENDER'),
                'receiver': header.get('RECEIVER'),
                'supplier': header.get('SUPPLIER'),
                'customer': header.get('CUSTOMER'),
                'profile_id': header.get('PROFILEID'),  # ✅ CREATE ile tutarlı
                'invoice_type_code': header.get('INVOICE_TYPE_CODE'),  # ✅ CREATE ile tutarlı
                'status': header.get('STATUS'),
                'status_description': header.get('STATUS_DESCRIPTION'),
                'status_code': header.get('STATUS_CODE'),
                'gib_status_code': header.get('GIB_STATUS_CODE'),
                'gib_status_description': header.get('GIB_STATUS_DESCRIPTION'),
                'response_code': header.get('RESPONSE_CODE'),  # YENİ (v1.0.4)
                'response_description': header.get('RESPONSE_DESCRIPTION'),  # YENİ (v1.0.4)
                'envelope_identifier': header.get('ENVELOPE_IDENTIFIER'),  # ✅ CREATE ile tutarlı
                'from_field': header.get('FROM'),  # ✅ CREATE ile tutarlı
                'to_field': header.get('TO'),  # ✅ CREATE ile tutarlı
            })

            # Tarihleri güncelle
            if header.get('ISSUE_DATE'):
                parsed_date = self._parse_date_field(header.get('ISSUE_DATE'), 'ISSUE_DATE')
                if parsed_date:
                    vals['issue_date'] = parsed_date

            if header.get('CDATE'):
                parsed_date = self._parse_date_field(header.get('CDATE'), 'CDATE')
                if parsed_date:
                    vals['create_date_ws'] = parsed_date

            # Finansal alanları güncelle
            financial_field_mapping = {
                'PAYABLE_AMOUNT': 'payable_amount',
                'TAX_EXCLUSIVE_TOTAL_AMOUNT': 'tax_exclusive_total_amount',
                'TAX_INCLUSIVE_TOTAL_AMOUNT': 'tax_inclusive_total_amount',
                'ALLOWANCE_TOTAL_AMOUNT': 'allowance_total_amount',
                'LINE_EXTENSION_AMOUNT': 'line_extension_amount'
            }

            for soap_field, odoo_field in financial_field_mapping.items():
                if header.get(soap_field):
                    parsed_amount = self._parse_financial_field(header.get(soap_field), soap_field)
                    vals[odoo_field] = parsed_amount

        return vals

    @api.model
    def sync_earsiv_from_soap(self, start_date, end_date):
        """
        E-Arşiv faturalarını SOAP servisinden çek ve Odoo'ya kaydet

        Args:
            start_date (str): Başlangıç tarihi (YYYY-MM-DD)
            end_date (str): Bitiş tarihi (YYYY-MM-DD)

        Returns:
            dict: {'success': bool, 'created': int, 'updated': int, 'message': str}
        """
        try:
            _logger.info("E-Arşiv senkronizasyonu başlatılıyor: %s - %s", start_date, end_date)

            # 1. Ortak SOAP client ve login
            efatura_client, session_id, transport_ws, settings_ws = self._get_soap_client_and_login()

            # 2. E-Arşiv Client oluştur
            soap_config = self.env['ir.config_parameter'].sudo()
            earsiv_ws = soap_config.get_param('efatura.earsiv_ws',
                'https://earsivws.izibiz.com.tr/EIArchiveWS/EFaturaArchive?wsdl')
            earsiv_client = Client(wsdl=earsiv_ws, transport=transport_ws, settings=settings_ws)

            # 3. GetEArchiveInvoiceList ile liste çek
            request_header = {'SESSION_ID': session_id}

            # Tarihleri datetime objesine çevir (SOAP servisi datetime bekliyor)
            from datetime import datetime as dt
            try:
                start_dt = dt.strptime(start_date, '%Y-%m-%d')
                end_dt = dt.strptime(end_date, '%Y-%m-%d')
            except:
                # Eğer parse edemezse bugünden başla
                end_dt = dt.now()
                start_dt = end_dt - timedelta(days=7)

            _logger.info("E-Arşiv fatura listesi çekiliyor: %s - %s", start_dt, end_dt)

            # E-Fatura pattern: raw_response ile XML parse
            with earsiv_client.settings(raw_response=True):
                raw_xml_response = earsiv_client.service.GetEArchiveInvoiceList(
                    REQUEST_HEADER=request_header,
                    LIMIT=25000,
                    START_DATE=start_dt,
                    END_DATE=end_dt,
                    HEADER_ONLY='Y',
                    READ_INCLUDED='true'
                )

            _logger.info("E-Arşiv SOAP Response alındı, XML parse ediliyor...")

            # 4. XML Parse (E-Fatura pattern with namespace fallback)
            root = ET.fromstring(raw_xml_response.content)
            invoices = []

            # Namespace fallback - namespace varsa da yoksa da çalışır
            for invoice_elem in root.findall('.//INVOICE') or root.findall('.//*[local-name()="INVOICE"]'):
                invoice_data = {}

                # Attribute'ler (ID, UUID, etc.)
                for attr_name, attr_value in invoice_elem.attrib.items():
                    invoice_data[attr_name] = attr_value

                # HEADER elementi - namespace fallback ile
                header_elem = invoice_elem.find('.//HEADER') or invoice_elem.find('.//*[local-name()="HEADER"]')
                if header_elem is not None:
                    header_data = {}
                    for child in header_elem:
                        tag_name = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                        header_data[tag_name] = child.text
                    invoice_data['HEADER'] = header_data

                invoices.append(invoice_data)

            _logger.info("Toplam %s E-Arşiv fatura bulundu", len(invoices))

            # 5. Logout
            efatura_client.service.Logout(REQUEST_HEADER=request_header)
            _logger.info("Logout başarılı")

            # 6. Response'u parse et ve kaydet
            created_count = 0
            updated_count = 0

            if invoices:
                for invoice_data in invoices:
                    try:
                        # Mevcut kayıt kontrolü (UUID ile - v1.0.6)
                        invoice_id = invoice_data.get('ID') or invoice_data.get('HEADER', {}).get('INVOICE_ID')
                        uuid = invoice_data.get('UUID') or invoice_data.get('HEADER', {}).get('UUID')

                        existing_invoice = self.search([
                            ('uuid', '=', uuid)
                        ], limit=1)

                        if existing_invoice:
                            # Kilitli kayıtları güncelleme (v1.0.5)
                            if existing_invoice.is_locked:
                                _logger.info(f"E-Arşiv Sync: Kilitli kayıt atlandı - {invoice_id}")
                                continue

                            # Güncelle
                            invoice_vals = self._prepare_earsiv_vals_from_soap(invoice_data)
                            existing_invoice.write(invoice_vals)
                            updated_count += 1
                            _logger.info("E-Arşiv fatura güncellendi: %s", invoice_id)
                        else:
                            # Yeni oluştur
                            self.create_from_earsiv_soap_data(invoice_data)
                            created_count += 1

                    except Exception as e:
                        _logger.error("E-Arşiv fatura kaydedilirken hata: %s", str(e))
                        continue
            else:
                _logger.info("E-Arşiv fatura bulunamadı")

            result = {
                'success': True,
                'created': created_count,
                'updated': updated_count,
                'message': _('%s yeni E-Arşiv fatura oluşturuldu, %s fatura güncellendi.') % (created_count, updated_count)
            }

            # Otomatik Logo senkronizasyonu kontrolü
            logo_auto_sync = soap_config.get_param('logo.auto_sync', False)

            if logo_auto_sync and result.get('success'):
                try:
                    if not pymssql:
                        _logger.warning("Otomatik Logo Sync (E-Arşiv): pymssql yüklü değil")
                        result['message'] += _('\n\nLogo Senkronizasyonu: pymssql eksik')
                    else:
                        # MSSQL config al
                        server = soap_config.get_param('logo.mssql_server')
                        port = int(soap_config.get_param('logo.mssql_port', '1433'))
                        database = soap_config.get_param('logo.mssql_database')
                        username = soap_config.get_param('logo.mssql_username')
                        password = soap_config.get_param('logo.mssql_password')
                        table_name = soap_config.get_param('logo.invoice_table', 'LG_600_01_INVOICE')

                        if not all([server, database, username, password]):
                            _logger.warning("Otomatik Logo Sync (E-Arşiv): MSSQL config eksik")
                            result['message'] += _('\n\nLogo Senkronizasyonu: MSSQL config eksik')
                        else:
                            # E-Arşiv faturaları al (her zaman OUT yönü)
                            domain = [
                                ('issue_date', '>=', start_date),
                                ('issue_date', '<=', end_date),
                                ('direction', '=', 'OUT'),
                                ('kaynak', '=', 'e-arsiv')
                            ]
                            invoices = self.search(domain)

                            if invoices:
                                # MSSQL bağlantısı
                                conn = pymssql.connect(
                                    server=server, port=port, user=username,
                                    password=password, database=database,
                                    timeout=30, login_timeout=30, charset='UTF-8'
                                )
                                cursor = conn.cursor()
                                stats = {'found': 0, 'not_found': 0, 'errors': 0}

                                # E-Arşiv her zaman OUT (giden)
                                trcode_condition = "TRCODE IN (6,7,8,9,14)"

                                # Her faturayı kontrol et
                                for invoice in invoices:
                                    try:
                                        # Tarih kontrolü için issue_date'i al
                                        invoice_date = invoice.issue_date.strftime('%Y-%m-%d') if invoice.issue_date else None

                                        if not invoice_date:
                                            _logger.warning("Logo sync E-Arşiv fatura %s: Tarih yok, atlanıyor", invoice.invoice_id)
                                            continue

                                        # Standart sorgu
                                        query = """
                                            SELECT LOGICALREF
                                            FROM {}
                                            WHERE (FICHENO = %s OR DOCODE = %s)
                                            AND CANCELLED = 0
                                            AND CAST(DATE_ AS DATE) = %s
                                            AND {}
                                        """.format(table_name, trcode_condition)

                                        cursor.execute(query, (invoice.invoice_id, invoice.invoice_id, invoice_date))
                                        logo_result = cursor.fetchone()

                                        if logo_result:
                                            invoice.write({
                                                'exists_in_logo': True,
                                                'logo_record_id': logo_result[0]
                                            })
                                            stats['found'] += 1
                                        else:
                                            invoice.write({'exists_in_logo': False})
                                            stats['not_found'] += 1
                                    except Exception as e:
                                        stats['errors'] += 1
                                        _logger.error("Logo sync E-Arşiv fatura %s: %s", invoice.invoice_id, str(e))

                                conn.close()
                                result['message'] += _('\n\nLogo Sync: %d bulundu, %d bulunamadı') % (stats['found'], stats['not_found'])
                            else:
                                result['message'] += _('\n\nLogo Sync: E-Arşiv fatura bulunamadı')

                except Exception as e:
                    import traceback
                    error_trace = traceback.format_exc()
                    _logger.warning("Otomatik Logo senkronizasyonu başarısız: %s\n%s", str(e), error_trace)
                    result['message'] += _('\n\nLogo Senkronizasyonu: Otomatik sync başarısız - %s') % str(e)

            return result

        except Exception as e:
            _logger.error("E-Arşiv SOAP senkronizasyon hatası: %s", str(e))
            return {
                'success': False,
                'error': str(e)
            }

    @api.model
    def create_from_earsiv_soap_data(self, soap_data):
        """
        E-Arşiv SOAP verisini Odoo modeline dönüştür ve kaydet

        WSDL Structure:
        - INVOICE
          - HEADER (dict with fields below)
            - INVOICE_ID, UUID, SENDER_NAME, SENDER_IDENTIFIER
            - CUSTOMER_NAME, CUSTOMER_IDENTIFIER
            - PROFILE_ID, INVOICE_TYPE, EARCHIVE_TYPE, SENDING_TYPE
            - STATUS, STATUS_CODE, ISSUE_DATE
            - PAYABLE_AMOUNT, TAXABLE_AMOUNT, CURRENCY_CODE
            - REPORTED

        Args:
            soap_data (dict): SOAP'tan gelen INVOICE data (with HEADER)

        Returns:
            recordset: Oluşturulan/Güncellenen e.invoice kaydı
        """
        # XML Parse sonucu gelen data yapısı:
        # soap_data = {
        #   'ID': '...',  # Attribute'den
        #   'UUID': '...', # Attribute'den
        #   'HEADER': {
        #     'INVOICE_ID': '...',
        #     'SENDER_NAME': '...',
        #     ...
        #   }
        # }

        header = soap_data.get('HEADER', {})

        invoice_vals = {
            # HARD-CODED Values
            'direction': 'OUT',  # E-Arşiv her zaman giden
            'kaynak': 'e-arsiv',
            'receiver': header.get('CUSTOMER_IDENTIFIER') or '4510016851',  # Müşteri VKN/TCKN

            # Ana Bilgiler - Attribute'lerden ve HEADER'dan
            'invoice_id': soap_data.get('ID') or header.get('INVOICE_ID'),
            'uuid': soap_data.get('UUID') or header.get('UUID'),
            'sender': header.get('SENDER_IDENTIFIER'),  # VKN/TCKN
            'supplier': header.get('SENDER_NAME'),  # Unvan
            'customer': header.get('CUSTOMER_NAME'),  # Müşteri adı
            'profile_id': header.get('PROFILE_ID'),  # EARSIVFATURA
            'invoice_type_code': header.get('INVOICE_TYPE'),  # SATIS, IADE
            'status': header.get('STATUS'),
            'status_code': header.get('STATUS_CODE'),

            # Tarih ve Tutarlar
            'issue_date': self._parse_date_field(header.get('ISSUE_DATE'), 'ISSUE_DATE'),
            'payable_amount': self._parse_financial_field(header.get('PAYABLE_AMOUNT'), 'PAYABLE_AMOUNT'),
            'tax_exclusive_total_amount': self._parse_financial_field(header.get('TAXABLE_AMOUNT'), 'TAXABLE_AMOUNT'),

            # E-Arşiv Özel Alanlar (4 yeni alan)
            'currency_code': header.get('CURRENCY_CODE'),
            'reported': header.get('REPORTED') == 'Y',
            'earchive_type': header.get('EARCHIVE_TYPE'),
            'sending_type': header.get('SENDING_TYPE'),
        }

        # Yeni kayıt oluştur
        new_record = self.create(invoice_vals)
        _logger.info("Yeni E-Arşiv fatura oluşturuldu: %s", invoice_vals['invoice_id'])
        return new_record

    def _prepare_earsiv_vals_from_soap(self, soap_data):
        """E-Arşiv SOAP verisini update vals formatına dönüştür"""
        header = soap_data.get('HEADER', {})

        vals = {
            # HARD-CODED Values (update sırasında değişmez ama tutarlılık için ekle)
            'direction': 'OUT',
            'kaynak': 'e-arsiv',
            'receiver': header.get('CUSTOMER_IDENTIFIER') or '4510016851',

            # Ana Bilgiler - Güncellenebilir
            'sender': header.get('SENDER_IDENTIFIER'),
            'supplier': header.get('SENDER_NAME'),
            'customer': header.get('CUSTOMER_NAME'),
            'profile_id': header.get('PROFILE_ID'),
            'invoice_type_code': header.get('INVOICE_TYPE'),
            'status': header.get('STATUS'),
            'status_code': header.get('STATUS_CODE'),

            # Tarih ve Tutarlar
            'issue_date': self._parse_date_field(header.get('ISSUE_DATE'), 'ISSUE_DATE'),
            'payable_amount': self._parse_financial_field(header.get('PAYABLE_AMOUNT'), 'PAYABLE_AMOUNT'),
            'tax_exclusive_total_amount': self._parse_financial_field(header.get('TAXABLE_AMOUNT'), 'TAXABLE_AMOUNT'),

            # E-Arşiv Özel Alanlar
            'currency_code': header.get('CURRENCY_CODE'),
            'reported': header.get('REPORTED') == 'Y',
            'earchive_type': header.get('EARCHIVE_TYPE'),
            'sending_type': header.get('SENDING_TYPE'),
        }

        return vals

    @api.model
    def _test_soap_connection(self):
        """SOAP servisi bağlantısını test et"""
        try:
            soap_config = self.env['ir.config_parameter'].sudo()
            username = soap_config.get_param('efatura.username')
            password = soap_config.get_param('efatura.password')
            
            if not username or not password:
                raise ValueError("Kimlik bilgileri eksik")
            
            session_ws = requests.Session()
            session_ws.verify = True
            transport_ws = Transport(session=session_ws)
            settings_ws = Settings(strict=False, xml_huge_tree=True, forbid_dtd=False, forbid_entities=False)
            
            efatura_ws = "https://efaturaws.izibiz.com.tr/EInvoiceWS?wsdl"
            efatura_client = Client(wsdl=efatura_ws, transport=transport_ws, settings=settings_ws)
            
            # Test login
            request_header = {
                'SESSION_ID': '-1',
                'APPLICATION_NAME': 'Odoo Test Client',
            }
            session_id = efatura_client.service.Login(
                REQUEST_HEADER=request_header,
                USER_NAME=username,
                PASSWORD=password
            ).SESSION_ID
            
            # Logout
            request_header['SESSION_ID'] = session_id
            efatura_client.service.Logout(REQUEST_HEADER=request_header)
            
            return session_id
            
        except Exception as e:
            _logger.error("SOAP bağlantı test hatası: %s", str(e))
            raise e

    # Cron Job Metodları
    @api.model
    def _check_working_hours(self, start_hour, end_hour):
        """Çalışma saatleri içinde mi kontrol et"""
        from datetime import datetime
        import pytz

        # Istanbul timezone'ında current hour'ı al
        istanbul_tz = pytz.timezone('Europe/Istanbul')
        current_time_istanbul = datetime.now(istanbul_tz)
        current_hour = current_time_istanbul.hour

        _logger.info(f"Çalışma saati kontrolü - Istanbul saati: {current_hour}, Başlangıç: {start_hour}, Bitiş: {end_hour}")

        if start_hour <= end_hour:
            return start_hour <= current_hour <= end_hour
        else:  # Gece yarısını geçen durumlar (ör: 22-02)
            return current_hour >= start_hour or current_hour <= end_hour

    @api.model
    def cron_progressive_sync(self):
        """Progressive Sync - 7 günlük periyotlarla e-fatura ve e-arşiv senkronizasyonu"""
        try:
            # Config parametrelerini al
            ICPSudo = self.env['ir.config_parameter'].sudo()

            # Enabled kontrolü
            enabled = ICPSudo.get_param('cron.progressive_sync_enabled', 'False') == 'True'
            if not enabled:
                return

            # Saat kontrolü
            start_hour = int(ICPSudo.get_param('cron.progressive_start_hour', '8'))
            end_hour = int(ICPSudo.get_param('cron.progressive_end_hour', '20'))
            if not self._check_working_hours(start_hour, end_hour):
                _logger.info("Progressive Sync: Çalışma saatleri dışında, atlanıyor...")
                return

            # Tarih parametrelerini al
            start_date_str = ICPSudo.get_param('cron.progressive_start_date')
            end_date_str = ICPSudo.get_param('cron.progressive_end_date')
            last_sync_date_str = ICPSudo.get_param('cron.progressive_last_sync_date')

            if not start_date_str or not end_date_str:
                _logger.warning("Progressive Sync: Başlangıç veya bitiş tarihi tanımlı değil")
                return

            # String'leri date objesine çevir
            start_date = fields.Date.from_string(start_date_str)
            end_date = fields.Date.from_string(end_date_str)

            # Son senkronize tarihi belirle
            if last_sync_date_str:
                current_date = fields.Date.from_string(last_sync_date_str)
                # 7 gün ekle
                from datetime import timedelta
                next_date = current_date + timedelta(days=7)

                # Bitiş tarihini geçtiyse başa dön
                if next_date > end_date:
                    _logger.info("Progressive Sync: Bitiş tarihine ulaşıldı, başa dönülüyor...")
                    current_date = start_date
                else:
                    current_date = next_date
            else:
                current_date = start_date

            # 7 günlük periyodu hesapla
            from datetime import timedelta
            period_end = min(current_date + timedelta(days=6), end_date)

            _logger.info("Progressive Sync başlatılıyor: %s - %s", current_date, period_end)

            # E-Fatura Gelen senkronizasyonu
            try:
                result_in = self.sync_invoices_from_soap(
                    current_date.strftime('%Y-%m-%d'),
                    period_end.strftime('%Y-%m-%d'),
                    'IN'
                )
                _logger.info("Progressive Sync - E-Fatura Gelen: %s", result_in)
            except Exception as e:
                _logger.error("Progressive Sync - E-Fatura Gelen hatası: %s", str(e))

            # E-Fatura Giden senkronizasyonu
            try:
                result_out = self.sync_invoices_from_soap(
                    current_date.strftime('%Y-%m-%d'),
                    period_end.strftime('%Y-%m-%d'),
                    'OUT'
                )
                _logger.info("Progressive Sync - E-Fatura Giden: %s", result_out)
            except Exception as e:
                _logger.error("Progressive Sync - E-Fatura Giden hatası: %s", str(e))

            # E-Arşiv senkronizasyonu
            try:
                result_earsiv = self.sync_earsiv_from_soap(
                    current_date.strftime('%Y-%m-%d'),
                    period_end.strftime('%Y-%m-%d')
                )
                _logger.info("Progressive Sync - E-Arşiv: %s", result_earsiv)
            except Exception as e:
                _logger.error("Progressive Sync - E-Arşiv hatası: %s", str(e))

            # Son senkronize tarihi güncelle (başlangıç tarihini kaydet, böylece +7 gün doğru çalışır)
            ICPSudo.set_param('cron.progressive_last_sync_date', current_date.strftime('%Y-%m-%d'))

            _logger.info("Progressive Sync tamamlandı. Sync edildi: %s - %s, Sonraki başlangıç: %s",
                        current_date, period_end, (current_date + timedelta(days=7)).strftime('%Y-%m-%d'))

        except Exception as e:
            _logger.error("Progressive Sync genel hatası: %s", str(e))

    @api.model
    def cron_retrospective_sync(self):
        """Retrospective Sync - X gün önceki tarihin faturalarını senkronize et"""
        try:
            # Config parametrelerini al
            ICPSudo = self.env['ir.config_parameter'].sudo()

            # Enabled kontrolü
            enabled = ICPSudo.get_param('cron.retrospective_sync_enabled', 'False') == 'True'
            if not enabled:
                return

            # Saat kontrolü
            start_hour = int(ICPSudo.get_param('cron.retrospective_start_hour', '8'))
            end_hour = int(ICPSudo.get_param('cron.retrospective_end_hour', '20'))
            if not self._check_working_hours(start_hour, end_hour):
                _logger.info("Retrospective Sync: Çalışma saatleri dışında, atlanıyor...")
                return

            # Kaç gün önce parametresi
            days_ago = int(ICPSudo.get_param('cron.retrospective_days_ago', '3'))

            # Hedef tarihi hesapla
            target_date = fields.Date.subtract(fields.Date.today(), days=days_ago)
            target_date_str = target_date.strftime('%Y-%m-%d')

            _logger.info("Retrospective Sync başlatılıyor: %s (%d gün önce)", target_date_str, days_ago)

            # E-Fatura Gelen senkronizasyonu
            try:
                result_in = self.sync_invoices_from_soap(
                    target_date_str,
                    target_date_str,
                    'IN'
                )
                _logger.info("Retrospective Sync - E-Fatura Gelen: %s", result_in)
            except Exception as e:
                _logger.error("Retrospective Sync - E-Fatura Gelen hatası: %s", str(e))

            # E-Fatura Giden senkronizasyonu
            try:
                result_out = self.sync_invoices_from_soap(
                    target_date_str,
                    target_date_str,
                    'OUT'
                )
                _logger.info("Retrospective Sync - E-Fatura Giden: %s", result_out)
            except Exception as e:
                _logger.error("Retrospective Sync - E-Fatura Giden hatası: %s", str(e))

            # E-Arşiv senkronizasyonu
            try:
                result_earsiv = self.sync_earsiv_from_soap(
                    target_date_str,
                    target_date_str
                )
                _logger.info("Retrospective Sync - E-Arşiv: %s", result_earsiv)
            except Exception as e:
                _logger.error("Retrospective Sync - E-Arşiv hatası: %s", str(e))

            _logger.info("Retrospective Sync tamamlandı: %s", target_date_str)

        except Exception as e:
            _logger.error("Retrospective Sync genel hatası: %s", str(e))

    @api.model
    def cron_logo_monthly_sync(self):
        """Logo Monthly Sync - 7 günlük periyotlarla Logo senkronizasyonu"""
        try:
            from datetime import timedelta

            # Config parametrelerini al
            ICPSudo = self.env['ir.config_parameter'].sudo()

            # Enabled kontrolü
            enabled = ICPSudo.get_param('cron.logo_monthly_sync_enabled', 'False') == 'True'
            if not enabled:
                return

            # Saat kontrolü
            start_hour = int(ICPSudo.get_param('cron.logo_monthly_start_hour', '8'))
            end_hour = int(ICPSudo.get_param('cron.logo_monthly_end_hour', '20'))
            if not self._check_working_hours(start_hour, end_hour):
                _logger.info("Logo Monthly Sync: Çalışma saatleri dışında, atlanıyor...")
                return

            # Tarih parametrelerini al
            start_date_str = ICPSudo.get_param('cron.logo_monthly_start_date')
            end_date_str = ICPSudo.get_param('cron.logo_monthly_end_date')
            last_sync_date_str = ICPSudo.get_param('cron.logo_monthly_last_sync_date')

            if not start_date_str or not end_date_str:
                _logger.warning("Logo Monthly Sync: Başlangıç veya bitiş tarihi tanımlı değil")
                return

            # String'leri date objesine çevir
            start_date = fields.Date.from_string(start_date_str)
            end_date = fields.Date.from_string(end_date_str)

            # Son senkronize tarihi belirle
            if last_sync_date_str:
                current_date = fields.Date.from_string(last_sync_date_str)
                # 7 gün ekle (bir sonraki periyoda geç)
                current_date = current_date + timedelta(days=7)

                # Bitiş tarihini geçtiyse başa dön
                if current_date > end_date:
                    _logger.info("Logo Monthly Sync: Bitiş tarihine ulaşıldı, başa dönülüyor...")
                    current_date = start_date
            else:
                current_date = start_date

            # 7 günlük periyot sonu hesapla
            period_end = current_date + timedelta(days=6)
            period_end = min(period_end, end_date)

            _logger.info("Logo Monthly Sync başlatılıyor: %s - %s", current_date, period_end)

            # Bu periyot için e_invoice kayıtlarını al (Kilitli kayıtları hariç tut - v1.0.5)
            domain = [
                ('issue_date', '>=', current_date),
                ('issue_date', '<=', period_end),
                ('kaynak', 'in', ['e-fatura', 'e-arsiv']),
                ('is_locked', '=', False)
            ]

            invoices = self.search(domain)

            if invoices:
                _logger.info("Logo Monthly Sync: %d fatura bulundu", len(invoices))

                # Logo sync işlemini doğrudan yap (wizard kullanmadan)
                try:
                    # MSSQL bağlantısını test et
                    if not pymssql:
                        _logger.warning("Logo Monthly Sync: pymssql kütüphanesi yüklü değil, atlanıyor...")
                    else:
                        # Config parametrelerini al
                        config_param = self.env['ir.config_parameter'].sudo()
                        server = config_param.get_param('logo.mssql_server')
                        port = int(config_param.get_param('logo.mssql_port', '1433'))
                        database = config_param.get_param('logo.mssql_database')
                        username = config_param.get_param('logo.mssql_username')
                        password = config_param.get_param('logo.mssql_password')
                        table_name = config_param.get_param('logo.invoice_table', 'LG_600_01_INVOICE')

                        if not all([server, database, username, password]):
                            _logger.warning("Logo Monthly Sync: MSSQL config eksik, atlanıyor...")
                        else:
                            # MSSQL bağlantısı aç
                            conn = pymssql.connect(
                                server=server,
                                port=port,
                                user=username,
                                password=password,
                                database=database,
                                timeout=30,
                                login_timeout=30,
                                charset='UTF-8'
                            )
                            cursor = conn.cursor()

                            # İstatistikler
                            stats = {'found': 0, 'not_found': 0, 'updated': 0, 'errors': 0}

                            # Her faturayı kontrol et
                            for invoice in invoices:
                                try:
                                    # Direction'a göre TRCODE belirle
                                    if invoice.direction == 'IN':
                                        trcode_condition = "TRCODE IN (1,3,4,13)"
                                    elif invoice.direction == 'OUT':
                                        trcode_condition = "TRCODE IN (6,7,8,9,14)"
                                    else:
                                        continue

                                    # Tarih kontrolü için issue_date'i al
                                    invoice_date = invoice.issue_date.strftime('%Y-%m-%d') if invoice.issue_date else None

                                    if not invoice_date:
                                        _logger.warning("Logo Monthly Sync - Fatura %s: Tarih yok, atlanıyor", invoice.invoice_id)
                                        continue

                                    # Standart sorgu
                                    query = """
                                        SELECT LOGICALREF
                                        FROM {}
                                        WHERE (FICHENO = %s OR DOCODE = %s)
                                        AND CANCELLED = 0
                                        AND CAST(DATE_ AS DATE) = %s
                                        AND {}
                                    """.format(table_name, trcode_condition)

                                    cursor.execute(query, (invoice.invoice_id, invoice.invoice_id, invoice_date))
                                    result = cursor.fetchone()

                                    if result:
                                        # Logo'da bulundu
                                        invoice.write({
                                            'exists_in_logo': True,
                                            'logo_record_id': result[0],
                                            'notes': (invoice.notes or '') + f'\n[Logo Sync] Bulundu: LOGICALREF={result[0]}'
                                        })
                                        stats['found'] += 1
                                    else:
                                        # Logo'da bulunamadı
                                        invoice.write({
                                            'exists_in_logo': False,
                                            'logo_record_id': False,
                                            'notes': (invoice.notes or '') + f'\n[Logo Sync] Bulunamadı'
                                        })
                                        stats['not_found'] += 1

                                    stats['updated'] += 1

                                except Exception as e:
                                    stats['errors'] += 1
                                    _logger.error("Logo Monthly Sync - Fatura %s hatası: %s", invoice.invoice_id, str(e))

                            conn.close()
                            _logger.info("Logo Monthly Sync tamamlandı: %d bulundu, %d bulunamadı, %d hata",
                                        stats['found'], stats['not_found'], stats['errors'])

                except Exception as e:
                    import traceback
                    error_trace = traceback.format_exc()
                    _logger.error("Logo Monthly Sync: Logo senkronizasyonu hatası: %s\n%s", str(e), error_trace)
            else:
                _logger.info("Logo Monthly Sync: Bu dönemde fatura bulunamadı")

            # Son senkronize tarihi güncelle (başlangıç tarihini kaydediyoruz, end değil)
            ICPSudo.set_param('cron.logo_monthly_last_sync_date', current_date.strftime('%Y-%m-%d'))

            _logger.info("Logo Monthly Sync tamamlandı. Sync edildi: %s - %s, Sonraki başlangıç: %s",
                        current_date, period_end, (current_date + timedelta(days=7)).strftime('%Y-%m-%d'))

        except Exception as e:
            _logger.error("Logo Monthly Sync genel hatası: %s", str(e))

    # Dashboard için metodlar
    @api.model
    def get_dashboard_data(self):
        """Dashboard için özet veriler"""
        domain_base = [('active', '=', True)]
        
        # Bu ayın verileri
        today = fields.Date.today()
        first_day_month = today.replace(day=1)
        
        current_month_domain = domain_base + [
            ('issue_date', '>=', first_day_month),
            ('issue_date', '<=', today)
        ]
        
        # Temel istatistikler
        total_invoices = self.search_count(domain_base)
        current_month_invoices = self.search_count(current_month_domain)                                                    
        incoming_invoices = self.search_count(domain_base + [('direction', '=', 'IN')])
        outgoing_invoices = self.search_count(domain_base + [('direction', '=', 'OUT')])
        
        # Finansal özet
        total_amount = sum(self.search(domain_base).mapped('payable_amount'))
        current_month_amount = sum(self.search(current_month_domain).mapped('payable_amount'))
        
        # Durum dağılımı
        status_data = self.read_group(
            domain_base,
            ['status'],
            ['status']
        )
        
        # Son 6 aylık trend
        monthly_data = self.read_group(
            domain_base + [('issue_date', '>=', fields.Date.subtract(today, months=6))],
            ['payable_amount', 'issue_date'],
            ['issue_date:month']
        )
        
        return {
            'total_invoices': total_invoices,
            'current_month_invoices': current_month_invoices,
            'incoming_invoices': incoming_invoices,
            'outgoing_invoices': outgoing_invoices,
            'total_amount': total_amount,
            'current_month_amount': current_month_amount,
            'status_distribution': status_data,
            'monthly_trend': monthly_data,
        }
    
    @api.model
    def get_top_customers_suppliers(self, limit=10):
        """En çok fatura gönderen/alan firmalar"""
        # En çok fatura gönderen (OUT direction)
        top_receivers = self.read_group(
            [('direction', '=', 'OUT'), ('active', '=', True), ('receiver', '!=', False)],
            ['receiver', 'payable_amount'],
            ['receiver'],
            limit=limit,
            orderby='payable_amount desc'
        )
        
        # En çok fatura gelen (IN direction)  
        top_senders = self.read_group(
            [('direction', '=', 'IN'), ('active', '=', True), ('sender', '!=', False)],
            ['sender', 'payable_amount'],
            ['sender'],
            limit=limit,
            orderby='payable_amount desc'
        )
        
        return {
            'top_receivers': top_receivers,
            'top_senders': top_senders
        }

    # Logo Sync metodları
    def action_open_logo_sync_wizard(self):
        """Seçili kayıtlar için Logo sync wizard'ını aç"""
        return {
            'type': 'ir.actions.act_window',
            'name': _('Logo ile Senkronize Et'),
            'res_model': 'logo.sync.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_sync_mode': 'selected',
                'active_ids': self.ids,
                'active_model': 'e.invoice',
            }
        }
    
    def action_sync_single_with_logo(self):
        """Tek kayıt için hızlı Logo sync"""
        self.ensure_one()
        
        # Wizard oluştur
        wizard = self.env['logo.sync.wizard'].create({
            'sync_mode': 'selected',
        })
        
        return {
            'type': 'ir.actions.act_window',
            'name': _('%s - Logo Senkronizasyonu') % self.invoice_id,
            'res_model': 'logo.sync.wizard',
            'res_id': wizard.id,
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_ids': [self.id],
                'active_model': 'e.invoice',
            }
        }


class EInvoiceReport(models.Model):
    _name = 'e.invoice.report'
    _description = 'E-Fatura Rapor Analizi'
    _auto = False
    _rec_name = 'date'

    date = fields.Date('Tarih')
    month = fields.Char('Ay')
    year = fields.Char('Yıl')
    direction = fields.Selection([('IN', 'Gelen'), ('OUT', 'Giden')], 'Yön')
    kaynak = fields.Selection([
        ('e-fatura', 'E-Fatura'),
        ('e-arsiv', 'E-Arşiv'),
        ('e-irsaliye', 'E-İrsaliye'),
    ], string='Kaynak')
    sender = fields.Char('Gönderen')
    receiver = fields.Char('Alıcı')
    invoice_count = fields.Integer('Fatura Sayısı')
    total_amount = fields.Float('Toplam Tutar')
    avg_amount = fields.Float('Ortalama Tutar')
    status = fields.Char('Durum')
    
    def init(self):
        tools.drop_view_if_exists(self.env.cr, self._table)
        self.env.cr.execute("""
            CREATE OR REPLACE VIEW %s AS (
                SELECT
                    row_number() OVER () AS id,
                    DATE(ei.issue_date) as date,
                    TO_CHAR(ei.issue_date, 'YYYY-MM') as month,
                    TO_CHAR(ei.issue_date, 'YYYY') as year,
                    ei.direction,
                    ei.kaynak,
                    ei.sender,
                    ei.receiver,
                    COUNT(*) as invoice_count,
                    SUM(ei.payable_amount) as total_amount,
                    AVG(ei.payable_amount) as avg_amount,
                    ei.status
                FROM e_invoice ei
                WHERE ei.active = true
                GROUP BY
                    DATE(ei.issue_date),
                    TO_CHAR(ei.issue_date, 'YYYY-MM'),
                    TO_CHAR(ei.issue_date, 'YYYY'),
                    ei.direction,
                    ei.kaynak,
                    ei.sender,
                    ei.receiver,
                    ei.status
            )
        """ % self._table)


class e_invoice_sync_wizard(models.TransientModel):
    _name = 'e.invoice.sync.wizard'
    _description = 'E-Fatura/E-Arşiv Senkronizasyon Sihirbazı'

    invoice_source = fields.Selection([
        ('e-fatura', 'E-Fatura'),
        ('e-arsiv', 'E-Arşiv'),
    ], string='Kaynak', default='e-fatura', required=True,
       help="Senkronize edilecek belge kaynağı")

    start_date = fields.Date(string='Başlangıç Tarihi', required=True)
    end_date = fields.Date(string='Bitiş Tarihi', required=True)
    direction = fields.Selection([
        ('IN', 'Gelen Faturalar'),
        ('OUT', 'Giden Faturalar')
    ], string='Yön', default='IN', required=True)
    
    def action_sync(self):
        """Senkronizasyon işlemini başlat"""

        # Tarih validasyonu
        if self.start_date > self.end_date:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'message': _("Hata: Başlangıç tarihi bitiş tarihinden büyük olamaz!"),
                    'type': 'danger',
                    'sticky': True,
                }
            }

        # Tarih aralığı kontrolü (en fazla 7 gün)
        date_diff = (self.end_date - self.start_date).days
        if date_diff > 7:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'message': _("Hata: En fazla 7 günlük tarih aralığı seçilebilir!"),
                    'type': 'danger',
                    'sticky': True,
                }
            }

        # Kaynak seçimine göre farklı metod çağır
        if self.invoice_source == 'e-fatura':
            # Mevcut E-Fatura senkronizasyonu
            result = self.env['e.invoice'].sync_invoices_from_soap(
                self.start_date.strftime('%Y-%m-%d'),
                self.end_date.strftime('%Y-%m-%d'),
                self.direction  # Kullanıcı seçimi
            )
        elif self.invoice_source == 'e-arsiv':
            # Yeni E-Arşiv senkronizasyonu
            result = self.env['e.invoice'].sync_earsiv_from_soap(
                self.start_date.strftime('%Y-%m-%d'),
                self.end_date.strftime('%Y-%m-%d')
                # direction parametresi YOK - kod içinde 'OUT' hard-coded
            )
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'message': _("Hata: Geçersiz kaynak seçimi!"),
                    'type': 'danger',
                    'sticky': True,
                }
            }

        if result.get('success'):
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'message': result.get('message'),
                    'type': 'success',
                    'sticky': False,
                }
            }
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'message': _("Hata: %s") % result.get('error'),
                    'type': 'danger',
                    'sticky': True,
                }
            }


class LogoSyncWizard(models.TransientModel):
    _name = 'logo.sync.wizard'
    _description = 'Logo MSSQL Senkronizasyon Sihirbazı'

    # Filtre Seçenekleri
    sync_mode = fields.Selection([
        ('all', 'Tüm Kayıtlar'),
        ('selected', 'Seçili Kayıtlar'),
        ('filtered', 'Filtrelenmiş Kayıtlar')
    ], string='Senkronizasyon Modu', default='all', required=True)
    
    date_filter = fields.Boolean(
        string='Tarih Filtresi Uygula',
        default=False
    )
    date_from = fields.Date(
        string='Başlangıç Tarihi',
        help="Bu tarihten sonraki faturalar senkronize edilir"
    )
    date_to = fields.Date(
        string='Bitiş Tarihi',
        help="Bu tarihten önceki faturalar senkronize edilir"
    )
    
    direction_filter = fields.Selection([
        ('all', 'Tümü'),
        ('IN', 'Sadece Gelen'),
        ('OUT', 'Sadece Giden')
    ], string='Yön Filtresi', default='all')
    
    # Test Modu
    test_mode = fields.Boolean(
        string='Test Modu',
        default=False,
        help="Test modunda sadece bağlantı test edilir, veri güncellenmez"
    )
    
    # Sonuç Bilgileri
    result_message = fields.Text(
        string='Sonuç',
        readonly=True
    )

    @api.onchange('date_filter')
    def _onchange_date_filter(self):
        """Tarih filtresi açıldığında varsayılan tarihler"""
        if self.date_filter and not self.date_from:
            self.date_from = fields.Date.subtract(fields.Date.today(), days=30)
        if self.date_filter and not self.date_to:
            self.date_to = fields.Date.today()

    def _get_mssql_connection(self):
        """MSSQL bağlantısı oluştur - Config parametrelerinden"""
        if not pymssql:
            raise UserError(_("pymssql kütüphanesi yüklü değil. 'pip install pymssql' komutunu çalıştırın."))
        
        # Config parametrelerinden bağlantı bilgilerini al
        config_param = self.env['ir.config_parameter'].sudo()
        server = config_param.get_param('logo.mssql_server')
        port = int(config_param.get_param('logo.mssql_port', '1433'))
        database = config_param.get_param('logo.mssql_database')
        username = config_param.get_param('logo.mssql_username')
        password = config_param.get_param('logo.mssql_password')
        
        # Zorunlu parametreleri kontrol et
        if not all([server, database, username, password]):
            missing_params = []
            if not server: missing_params.append('Server')
            if not database: missing_params.append('Database')
            if not username: missing_params.append('Username')
            if not password: missing_params.append('Password')
            
            raise UserError(_("Logo MSSQL bağlantı parametreleri eksik: %s\n\nLütfen E-Fatura → Yapılandırma menüsünden ayarları tamamlayın.") % ', '.join(missing_params))
        
        try:
            connection = pymssql.connect(
                server=server,
                port=port,
                user=username,
                password=password,
                database=database,
                timeout=30,
                login_timeout=30,
                charset='UTF-8'
            )
            return connection
        except Exception as e:
            raise UserError(_("MSSQL bağlantı hatası: %s\n\nBağlantı ayarlarını kontrol edin: %s:%s@%s/%s") % (str(e), username, port, server, database))

    def action_test_connection(self):
        """MSSQL bağlantısını test et"""
        try:
            conn = self._get_mssql_connection()
            cursor = conn.cursor()
            
            # Config'den tablo adını al
            config_param = self.env['ir.config_parameter'].sudo()
            table_name = config_param.get_param('logo.invoice_table', 'LG_600_01_INVOICE')
            
            # Test sorgusu çalıştır
            cursor.execute("SELECT COUNT(*) FROM {}".format(table_name))
            result = cursor.fetchone()
            invoice_count = result[0] if result else 0
            
            conn.close()
            
            message = _("✅ Bağlantı başarılı!\n\nLogo veritabanında %s fatura kaydı bulundu.") % "{:,}".format(invoice_count)
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Bağlantı Testi Başarılı'),
                    'message': message,
                    'type': 'success',
                    'sticky': False,
                }
            }
            
        except Exception as e:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Bağlantı Hatası'),
                    'message': _("❌ %s") % str(e),
                    'type': 'danger',
                    'sticky': True,
                }
            }

    def _get_e_invoices_to_sync(self):
        """Senkronize edilecek e-fatura kayıtlarını getir"""
        domain = []
        
        # Yön filtresi
        if self.direction_filter != 'all':
            domain.append(('direction', '=', self.direction_filter))
        
        # Tarih filtresi
        if self.date_filter:
            if self.date_from:
                domain.append(('issue_date', '>=', self.date_from))
            if self.date_to:
                domain.append(('issue_date', '<=', self.date_to))
        
        # Senkronizasyon modu
        if self.sync_mode == 'selected':
            # Context'ten seçili ID'leri al
            active_ids = self.env.context.get('active_ids', [])
            if active_ids:
                domain.append(('id', 'in', active_ids))
            else:
                raise UserError(_("Seçili kayıt bulunamadı!"))
        
        return self.env['e.invoice'].search(domain)

    def _check_invoice_in_logo(self, cursor, invoice_id, direction, invoice_date):
        """Tek bir faturanın Logo'daki durumunu kontrol et"""
        try:
            # Config'den tablo adını al
            config_param = self.env['ir.config_parameter'].sudo()
            table_name = config_param.get_param('logo.invoice_table', 'LG_600_01_INVOICE')

            # Direction'a göre TRCODE değerlerini belirle
            if direction == 'IN':
                trcode_condition = "TRCODE IN (1,3,4,13)"
            elif direction == 'OUT':
                trcode_condition = "TRCODE IN (6,7,8,9,14)"
            else:
                return {
                    'exists': False,
                    'logo_record_id': None,
                    'note': _('Geçersiz direction değeri: %s') % direction
                }

            # Datetime'ı sadece tarih kısmına çevir (saat bilgisini kaldır)
            # invoice_date datetime ise date'e çevir, yoksa olduğu gibi kullan
            if invoice_date:
                if isinstance(invoice_date, datetime):
                    date_only = invoice_date.date()
                else:
                    date_only = invoice_date
            else:
                # Tarih yoksa sorguyu yapma
                return {
                    'exists': False,
                    'logo_record_id': None,
                    'note': _('Fatura tarihi bulunamadı')
                }

            # SQL sorgusu - CAST ile datetime'ı date'e çeviriyoruz
            query = """
                SELECT LOGICALREF
                FROM {}
                WHERE (FICHENO = %s OR DOCODE = %s)
                AND CANCELLED = 0
                AND CAST(DATE_ AS DATE) = %s
                AND {}
            """.format(table_name, trcode_condition)

            cursor.execute(query, (invoice_id, invoice_id, date_only))
            results = cursor.fetchall()
            
            if len(results) == 0:
                return {
                    'exists': False,
                    'logo_record_id': None,
                    'note': _('Logo veri tabanında bu fatura kaydı bulunamadı')
                }
            elif len(results) == 1:
                return {
                    'exists': True,
                    'logo_record_id': results[0][0],
                    'note': _('Logo eşi var')
                }
            else:
                return {
                    'exists': False,
                    'logo_record_id': None,
                    'note': _('Logo veri tabanında birden fazla eş kayıt bulundu')
                }
                
        except Exception as e:
            return {
                'exists': False,
                'logo_record_id': None,
                'note': _('Logo sorgu hatası: %s') % str(e)
            }

    def action_sync_logo(self):
        """Ana senkronizasyon işlemi"""
        
        try:
            # E-fatura kayıtlarını getir
            e_invoices = self._get_e_invoices_to_sync()
            
            if not e_invoices:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': _('Uyarı'),
                        'message': _('Senkronize edilecek fatura bulunamadı!'),
                        'type': 'warning',
                        'sticky': False,
                    }
                }
            
            # Test modu kontrolü
            if self.test_mode:
                return self._run_test_mode(e_invoices)
            
            # MSSQL bağlantısı
            conn = self._get_mssql_connection()
            cursor = conn.cursor()
            
            # Senkronizasyon istatistikleri
            stats = {
                'total': len(e_invoices),
                'found': 0,
                'not_found': 0,
                'multiple': 0,
                'errors': 0,
                'updated': 0
            }
            
            error_details = []
            
            # Her faturayı işle
            for invoice in e_invoices:
                try:
                    # Kilitli kayıtları atlama (v1.0.5)
                    if invoice.is_locked:
                        _logger.info(f"Logo Sync: Kilitli kayıt atlandı - {invoice.invoice_id}")
                        continue

                    # Logo'da kontrol et
                    logo_result = self._check_invoice_in_logo(
                        cursor,
                        invoice.invoice_id,
                        invoice.direction,
                        invoice.issue_date
                    )

                    # Mevcut notları koru
                    existing_notes = invoice.notes or ''
                    new_note = logo_result['note']

                    if existing_notes:
                        updated_notes = "{}\n{}".format(existing_notes, new_note)
                    else:
                        updated_notes = new_note

                    # E-fatura kaydını güncelle
                    update_vals = {
                        'exists_in_logo': logo_result['exists'],
                        'logo_record_id': logo_result['logo_record_id'],
                        'notes': updated_notes
                    }

                    invoice.write(update_vals)
                    stats['updated'] += 1
                    
                    # İstatistikleri güncelle
                    if logo_result['exists']:
                        stats['found'] += 1
                    elif 'bulunamadı' in logo_result['note']:
                        stats['not_found'] += 1
                    elif 'birden fazla' in logo_result['note']:
                        stats['multiple'] += 1
                    
                except Exception as e:
                    stats['errors'] += 1
                    error_msg = _("Fatura %s: %s") % (invoice.invoice_id, str(e))
                    error_details.append(error_msg)
                    _logger.error("Logo senkronizasyon hatası - %s", error_msg)
            
            conn.close()
            
            # Sonuç mesajı oluştur
            result_message = self._create_result_message(stats, error_details)
            
            # Wizard'ı güncelle
            self.write({'result_message': result_message})
            
            # Bildirim göster
            notification_type = 'success' if stats['errors'] == 0 else 'warning'
            notification_message = _("%s kayıt güncellendi, %s hata oluştu.") % (stats['updated'], stats['errors'])
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Senkronizasyon Tamamlandı'),
                    'message': notification_message,
                    'type': notification_type,
                    'sticky': False,
                }
            }
            
        except Exception as e:
            error_msg = _("Senkronizasyon hatası: %s") % str(e)
            _logger.error(error_msg)
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Kritik Hata'),
                    'message': error_msg,
                    'type': 'danger',
                    'sticky': True,
                }
            }

    def _run_test_mode(self, e_invoices):
        """Test modu çalıştır"""
        try:
            conn = self._get_mssql_connection()
            cursor = conn.cursor()
            
            # İlk 5 kaydı test et
            test_invoices = e_invoices[:5]
            test_results = []
            
            for invoice in test_invoices:
                logo_result = self._check_invoice_in_logo(
                    cursor,
                    invoice.invoice_id,
                    invoice.direction,
                    invoice.issue_date
                )

                test_results.append({
                    'invoice_id': invoice.invoice_id,
                    'direction': invoice.direction,
                    'issue_date': invoice.issue_date,
                    'exists': logo_result['exists'],
                    'logo_id': logo_result['logo_record_id'],
                    'note': logo_result['note']
                })
            
            conn.close()
            
            # Test sonuçları mesajı
            test_message = _("🔍 TEST MODU SONUÇLARI\n\n")
            test_message += _("Toplam kayıt sayısı: %s\n") % len(e_invoices)
            test_message += _("Test edilen kayıt sayısı: %s\n\n") % len(test_results)
            
            for result in test_results:
                status = _("✅ Bulundu") if result['exists'] else _("❌ Bulunamadı/Hata")
                test_message += _("Fatura: %s (%s) - %s\n") % (result['invoice_id'], result['direction'], status)
                test_message += _("   Not: %s\n\n") % result['note']
            
            self.write({'result_message': test_message})
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Test Tamamlandı'),
                    'message': _('%s kayıt test edildi. Detaylar için wizard penceresini kontrol edin.') % len(test_results),
                    'type': 'info',
                    'sticky': False,
                }
            }
            
        except Exception as e:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Test Hatası'),
                    'message': _("Test modu hatası: %s") % str(e),
                    'type': 'danger',
                    'sticky': True,
                }
            }

    def _create_result_message(self, stats, error_details):
        """Sonuç mesajı oluştur"""
        message = _("📊 LOGO SENKRONIZASYON SONUÇLARI\n\n")
        message += _("Toplam işlenen kayıt: %s\n") % stats['total']
        message += _("Güncellenen kayıt: %s\n\n") % stats['updated']
        
        message += _("📋 DETAY İSTATİSTİKLER:\n")
        message += _("✅ Logo'da bulunan: %s\n") % stats['found']
        message += _("❌ Logo'da bulunamayan: %s\n") % stats['not_found']
        message += _("⚠️ Birden fazla eş bulunan: %s\n") % stats['multiple']
        message += _("🚫 Hata oluşan: %s\n\n") % stats['errors']
        
        if error_details:
            message += _("🚨 HATA DETAYLARI:\n")
            for error in error_details[:10]:  # İlk 10 hatayı göster
                message += "• {}\n".format(error)
            
            if len(error_details) > 10:
                message += _("... ve %s hata daha\n") % (len(error_details) - 10)
        
        return message


class EInvoiceConfigSettings(models.TransientModel):
    _inherit = 'res.config.settings'
    
    # E-Fatura SOAP Ayarları
    efatura_username = fields.Char(
        string='E-Fatura Kullanıcı Adı',
        config_parameter='efatura.username'
    )
    efatura_password = fields.Char(
        string='E-Fatura Şifre',
        config_parameter='efatura.password'
    )
    efatura_earsiv_ws = fields.Char(
        string='E-Arşiv WS URL',
        config_parameter='efatura.earsiv_ws',
        default='https://earsivws.izibiz.com.tr/EIArchiveWS/EFaturaArchive?wsdl',
        help="E-Arşiv fatura SOAP servis adresi"
    )

    # Logo MSSQL Ayarları
    logo_mssql_server = fields.Char(
        string='MSSQL Server',
        config_parameter='logo.mssql_server',
        help="MSSQL Server adı veya IP adresi (örn: localhost, 192.168.1.100)"
    )
    logo_mssql_port = fields.Integer(
        string='MSSQL Port',
        config_parameter='logo.mssql_port',
        default=1433,
        help="MSSQL Server portu (varsayılan: 1433)"
    )
    logo_mssql_database = fields.Char(
        string='Logo Veritabanı',
        config_parameter='logo.mssql_database',
        help="Logo veritabanı adı (örn: TIGER_DB, LOGO_01)"
    )
    logo_mssql_username = fields.Char(
        string='MSSQL Kullanıcı Adı',
        config_parameter='logo.mssql_username',
        help="MSSQL veritabanı kullanıcı adı"
    )
    logo_mssql_password = fields.Char(
        string='MSSQL Şifre',
        config_parameter='logo.mssql_password',
        help="MSSQL veritabanı şifresi"
    )
    logo_invoice_table = fields.Char(
        string='Fatura Tablosu',
        config_parameter='logo.invoice_table',
        default='LG_600_01_INVOICE',
        help="Logo fatura tablosu adı (örn: LG_600_01_INVOICE, LG_001_01_INVOICE)"
    )
    logo_auto_sync = fields.Boolean(
        string='Logo Otomatik Senkronizasyon',
        config_parameter='logo.auto_sync',
        help="E-fatura senkronizasyonu sonrası otomatik olarak Logo senkronizasyonu çalıştır"
    )

    # Cron 1: Progressive Sync (7 Günlük Periyot)
    cron1_enabled = fields.Boolean(
        string='Progressive Sync Aktif',
        config_parameter='cron.progressive_sync_enabled',
        help="7 günlük periyotlarla e-fatura ve e-arşiv senkronizasyonu"
    )
    cron1_start_date = fields.Char(
        string='Başlangıç Tarihi',
        config_parameter='cron.progressive_start_date',
        help="YYYY-MM-DD formatında tarih girin"
    )
    cron1_end_date = fields.Char(
        string='Bitiş Tarihi',
        config_parameter='cron.progressive_end_date',
        help="YYYY-MM-DD formatında tarih girin"
    )
    cron1_last_sync_date = fields.Char(
        string='Son Senkronize Tarih',
        config_parameter='cron.progressive_last_sync_date',
        readonly=True
    )
    cron1_start_hour = fields.Integer(
        string='Başlangıç Saati',
        config_parameter='cron.progressive_start_hour',
        default=8,
        help="0-23 arası saat (örn: 8 = 08:00)"
    )
    cron1_end_hour = fields.Integer(
        string='Bitiş Saati',
        config_parameter='cron.progressive_end_hour',
        default=20,
        help="0-23 arası saat (örn: 20 = 20:00)"
    )

    # Cron 2: Retrospective Daily Sync (X Gün Önce)
    cron2_enabled = fields.Boolean(
        string='Retrospective Sync Aktif',
        config_parameter='cron.retrospective_sync_enabled',
        help="Belirtilen gün sayısı kadar önceki günün faturalarını senkronize et"
    )
    cron2_days_ago = fields.Integer(
        string='Kaç Gün Önce',
        config_parameter='cron.retrospective_days_ago',
        default=3,
        help="Bugünden kaç gün öncesinin faturalarını çekecek"
    )
    cron2_start_hour = fields.Integer(
        string='Başlangıç Saati',
        config_parameter='cron.retrospective_start_hour',
        default=8,
        help="0-23 arası saat (örn: 8 = 08:00)"
    )
    cron2_end_hour = fields.Integer(
        string='Bitiş Saati',
        config_parameter='cron.retrospective_end_hour',
        default=20,
        help="0-23 arası saat (örn: 20 = 20:00)"
    )

    # Cron 3: Logo Monthly Sync (7 Günlük Periyotlar)
    cron3_enabled = fields.Boolean(
        string='Logo Monthly Sync Aktif',
        config_parameter='cron.logo_monthly_sync_enabled',
        help="7 günlük periyotlarla Logo senkronizasyonu"
    )
    cron3_start_date = fields.Char(
        string='Başlangıç Tarihi',
        config_parameter='cron.logo_monthly_start_date',
        help="YYYY-MM-DD formatında tarih girin"
    )
    cron3_end_date = fields.Char(
        string='Bitiş Tarihi',
        config_parameter='cron.logo_monthly_end_date',
        help="YYYY-MM-DD formatında tarih girin"
    )
    cron3_last_sync_date = fields.Char(
        string='Son Senkronize Tarih',
        config_parameter='cron.logo_monthly_last_sync_date',
        readonly=True
    )
    cron3_start_hour = fields.Integer(
        string='Başlangıç Saati',
        config_parameter='cron.logo_monthly_start_hour',
        default=8,
        help="0-23 arası saat (örn: 8 = 08:00)"
    )
    cron3_end_hour = fields.Integer(
        string='Bitiş Saati',
        config_parameter='cron.logo_monthly_end_hour',
        default=20,
        help="0-23 arası saat (örn: 20 = 20:00)"
    )

    @api.model
    def get_values(self):
        res = super(EInvoiceConfigSettings, self).get_values()
        ICPSudo = self.env['ir.config_parameter'].sudo()
        
        # Boolean string'den bool'a çevirme helper
        def str_to_bool(val):
            if isinstance(val, str):
                return val.lower() in ('true', '1', 'yes', 'on')
            return bool(val) if val is not None else False
        
        res.update(
            # E-Fatura ayarları
            efatura_username=ICPSudo.get_param('efatura.username', ''),
            efatura_password=ICPSudo.get_param('efatura.password', ''),
            efatura_earsiv_ws=ICPSudo.get_param('efatura.earsiv_ws',
                'https://earsivws.izibiz.com.tr/EIArchiveWS/EFaturaArchive?wsdl'),

            # Logo MSSQL ayarları
            logo_mssql_server=ICPSudo.get_param('logo.mssql_server', ''),
            logo_mssql_port=int(ICPSudo.get_param('logo.mssql_port', '1433')),
            logo_mssql_database=ICPSudo.get_param('logo.mssql_database', ''),
            logo_mssql_username=ICPSudo.get_param('logo.mssql_username', ''),
            logo_mssql_password=ICPSudo.get_param('logo.mssql_password', ''),
            logo_invoice_table=ICPSudo.get_param('logo.invoice_table', 'LG_600_01_INVOICE'),
            logo_auto_sync=str_to_bool(ICPSudo.get_param('logo.auto_sync', 'False')),  # ✅ Düzeltildi
        )
        return res

    def set_values(self):
        super(EInvoiceConfigSettings, self).set_values()
        ICPSudo = self.env['ir.config_parameter'].sudo()
        
        # E-Fatura ayarları
        ICPSudo.set_param('efatura.username', self.efatura_username or '')
        ICPSudo.set_param('efatura.password', self.efatura_password or '')
        ICPSudo.set_param('efatura.earsiv_ws', self.efatura_earsiv_ws or
            'https://earsivws.izibiz.com.tr/EIArchiveWS/EFaturaArchive?wsdl')
        
        # Logo MSSQL ayarları
        ICPSudo.set_param('logo.mssql_server', self.logo_mssql_server or '')
        ICPSudo.set_param('logo.mssql_port', str(self.logo_mssql_port or 1433))  # ✅ String'e çevir
        ICPSudo.set_param('logo.mssql_database', self.logo_mssql_database or '')
        ICPSudo.set_param('logo.mssql_username', self.logo_mssql_username or '')
        ICPSudo.set_param('logo.mssql_password', self.logo_mssql_password or '')
        ICPSudo.set_param('logo.invoice_table', self.logo_invoice_table or 'LG_600_01_INVOICE')  # ✅ Eksik satır eklendi
        ICPSudo.set_param('logo.auto_sync', str(self.logo_auto_sync))  # ✅ String'e çevir

    def action_test_efatura_connection(self):
        """E-Fatura SOAP servisi bağlantısını test et"""
        try:
            session_id = self.env['e.invoice']._test_soap_connection()
            if session_id:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': _('E-Fatura Bağlantı Testi'),
                        'message': _('SOAP servisi bağlantısı başarılı!'),
                        'type': 'success',
                        'sticky': False,
                    }
                }
        except Exception as e:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('E-Fatura Bağlantı Hatası'),
                    'message': _('Bağlantı hatası: %s') % str(e),
                    'type': 'danger',
                    'sticky': True,
                }
            }
    
    def action_test_logo_connection(self):
        """Logo MSSQL bağlantısını test et"""
        try:
            # Geçici wizard oluştur ve bağlantı test et
            wizard = self.env['logo.sync.wizard'].create({})
            return wizard.action_test_connection()

        except Exception as e:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Logo Bağlantı Hatası'),
                    'message': _('Bağlantı hatası: %s') % str(e),
                    'type': 'danger',
                    'sticky': True,
                }
            }

    def action_reset_progressive_sync(self):
        """Progressive sync last_sync_date'i sıfırla"""
        ICPSudo = self.env['ir.config_parameter'].sudo()
        ICPSudo.set_param('cron.progressive_last_sync_date', False)

        # Form alanını da güncelle
        self.cron1_last_sync_date = False

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Progressive Sync Sıfırlandı'),
                'message': _('Son senkronize tarih sıfırlandı. Bir sonraki çalışmada başlangıç tarihinden başlayacak.'),
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }

    def action_reset_logo_monthly_sync(self):
        """Logo monthly sync last_sync_date'i sıfırla"""
        ICPSudo = self.env['ir.config_parameter'].sudo()
        ICPSudo.set_param('cron.logo_monthly_last_sync_date', False)

        # Form alanını da güncelle
        self.cron3_last_sync_date = False

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Logo Monthly Sync Sıfırlandı'),
                'message': _('Son senkronize tarih sıfırlandı. Bir sonraki çalışmada başlangıç tarihinden başlayacak.'),
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }


class LogoKdv2Report(models.TransientModel):
    _name = 'logo.kdv2.report'
    _description = 'Logo KDV-2 Raporu'
    
    # Rapor sonuçları için alanlar
    logo_id = fields.Integer(string='Logo ID', readonly=True)
    ay = fields.Integer(string='Ay', readonly=True)
    yil = fields.Integer(string='Yıl', readonly=True)
    fis_no = fields.Char(string='Fiş No', readonly=True)
    proje = fields.Char(string='Proje', readonly=True)
    kebir_hesap_kodu = fields.Char(string='Kebir Hesap Kodu', readonly=True)
    kebir_hesap_adi = fields.Char(string='Kebir Hesap Adı', readonly=True)
    hesap_kodu = fields.Char(string='Hesap Kodu', readonly=True)
    hesap_adi = fields.Char(string='Hesap Adı', readonly=True)
    masraf_merkezi = fields.Char(string='Masraf Merkezi', readonly=True)
    kaynak_modul = fields.Char(string='Kaynak Modül', readonly=True)
    aciklama = fields.Char(string='Açıklama', readonly=True)
    fis_aciklama = fields.Char(string='Fiş Açıklama', readonly=True)
    cari = fields.Char(string='Cari', readonly=True)
    cari_vergi_no = fields.Char(string='Cari Vergi No', readonly=True)
    cari_unvan = fields.Char(string='Cari Ünvan', readonly=True)
    adi = fields.Char(string='Adı', readonly=True)
    soy_adi = fields.Char(string='Soyadı', readonly=True)
    tckn = fields.Char(string='TCKN', readonly=True)
    tutar_yerel = fields.Float(string='Tutar (Yerel)', digits=(16, 2), readonly=True)
    kdv_tutar = fields.Float(string='KDV Tutar', digits=(16, 2), readonly=True)
    tevkifat_oran = fields.Char(string='Tevkifat Oran', readonly=True)
    tevkif_edilen_kdv_tutari = fields.Float(string='Tevkif Edilen KDV Tutarı', digits=(16, 2), readonly=True)


class LogoKdv2Wizard(models.TransientModel):
    _name = 'logo.kdv2.wizard'
    _description = 'Logo KDV-2 Rapor Sihirbazı'
    
    month = fields.Selection([
        ('1', 'Ocak'),
        ('2', 'Şubat'),
        ('3', 'Mart'),
        ('4', 'Nisan'),
        ('5', 'Mayıs'),
        ('6', 'Haziran'),
        ('7', 'Temmuz'),
        ('8', 'Ağustos'),
        ('9', 'Eylül'),
        ('10', 'Ekim'),
        ('11', 'Kasım'),
        ('12', 'Aralık'),
    ], string='Ay', required=True, default=str(fields.Date.today().month))
    
    year = fields.Integer(string='Yıl', required=True, default=fields.Date.today().year)
    
    def _get_mssql_connection(self):
        """MSSQL bağlantısı oluştur"""
        if not pymssql:
            raise UserError(_("pymssql kütüphanesi yüklü değil. 'pip install pymssql' komutunu çalıştırın."))
        
        config_param = self.env['ir.config_parameter'].sudo()
        server = config_param.get_param('logo.mssql_server')
        port = int(config_param.get_param('logo.mssql_port', '1433'))
        database = config_param.get_param('logo.mssql_database')
        username = config_param.get_param('logo.mssql_username')
        password = config_param.get_param('logo.mssql_password')
        
        if not all([server, database, username, password]):
            raise UserError(_("Logo MSSQL bağlantı parametreleri eksik. Lütfen E-Fatura → Yapılandırma menüsünden ayarları tamamlayın."))
        
        try:
            connection = pymssql.connect(
                server=server,
                port=str(port),
                user=username,
                password=password,
                database=database,
                timeout=120,
                login_timeout=60,
                charset='UTF-8',
                appname='Odoo E-Fatura'
            )
            return connection
        except Exception as e:
            raise UserError(_("MSSQL bağlantı hatası: %s") % str(e))
    
    def action_generate_report(self):
        """KDV-2 raporunu oluştur"""
        conn = None
        cursor = None
        try:
            # Mevcut kayıtları temizle
            self.env['logo.kdv2.report'].search([]).unlink()

            # MSSQL bağlantısı
            conn = self._get_mssql_connection()
            cursor = conn.cursor(as_dict=True)
            
            # SQL sorgusunu çalıştır
            query = """
SELECT  
 AA.LOGICALREF as logoID,
 MONTH(A.DATE_) as ay,
 YEAR(A.DATE_) as yil,
 AA.FICHENO as fisNo,
 E.CODE+' '+E.NAME as proje,
 F1.CODE as kebirHesapKodu,
 F1.DEFINITION_ as kebirHesapAdi,
 F.CODE as hesapKodu,
 F.DEFINITION_ as hesapAdi,
 G.CODE+' '+G.DEFINITION_ as masrafMerkezi,
 CASE WHEN AA.MODULENR=1 THEN '1 Malzeme'
  WHEN AA.MODULENR=2 THEN '2 Satınalma'
  WHEN AA.MODULENR=3 THEN '3 Satış'
  WHEN AA.MODULENR=4 THEN '4 Cari Hesap'
  WHEN AA.MODULENR=5 THEN '5 Çek Senet'
  WHEN AA.MODULENR=6 THEN '6 Banka'
  WHEN AA.MODULENR=7 THEN '7 Kasa'
  ELSE '' END as kaynakModul,
 A.LINEEXP as aciklama,
 AA.GENEXP1 as fisAciklama,
 A.CLDEF as cari, 
 A.TAXNR as cariVergiNo,
 CL.DEFINITION_ as cariUnvan, 
 CL.NAME as adi, 
 CL.SURNAME as soyAdi, 
 CL.TCKNO as tckn,
 CASE WHEN A.LOGICALREF NOT IN (SELECT DISTINCT PREVLINEREF FROM LG_600_01_ACCDISTDETLN) OR A1.DISTRATE=1 THEN ABS(A.DEBIT-A.CREDIT)-ABS(A.DEBIT-A.CREDIT)*2*A.SIGN ELSE A1.CREDEBNET-A1.CREDEBNET*2*A.SIGN END as tutarYerel,
 SEML.VATAMOUNT as kdvTutar,
 SEML.DEDUCTION as tevkifatOran,
 CASE WHEN ISNULL(SSTL.DEDUCTIONPART1,0) != 0 
                                    AND ISNULL(SSTL.DEDUCTIONPART2,0) != 0 
                                    AND ISNULL(SSTL.VAT,0) != 0
                                            THEN  ROUND((SSTL.GROSSTOTAL*SSTL.VAT/100) *  CAST(SSTL.DEDUCTIONPART1  AS FLOAT)  / CAST(SSTL.DEDUCTIONPART2 AS FLOAT),2)  ELSE 0 END as tevkifEdilenKdvTutari
FROM  LG_600_01_EMFLINE A WITH(NOLOCK)
 LEFT JOIN LG_600_01_EMFICHE AA WITH(NOLOCK) ON AA.LOGICALREF=A.ACCFICHEREF
 LEFT JOIN LG_600_01_ACCDISTDETLN A1 WITH(NOLOCK) ON A1.PREVLINEREF=A.LOGICALREF
 LEFT JOIN L_CAPIDIV C WITH(NOLOCK) ON C.NR=A.BRANCH AND C.FIRMNR=600
 LEFT JOIN L_CAPIDEPT D WITH(NOLOCK) ON D.NR=A.DEPARTMENT AND D.FIRMNR=600
 LEFT JOIN LG_600_PROJECT E WITH(NOLOCK) ON E.LOGICALREF=A1.PROJECTREF
 LEFT JOIN LG_600_EMUHACC F WITH(NOLOCK) ON F.LOGICALREF=A.ACCOUNTREF
 LEFT JOIN LG_600_EMUHACC F1 WITH(NOLOCK) ON F1.CODE=left(F.CODE,3)
 LEFT JOIN LG_600_EMCENTER G WITH(NOLOCK) ON G.LOGICALREF=A.CENTERREF
 LEFT JOIN L_CURRENCYLIST H WITH(NOLOCK) ON H.CURTYPE=A.TRCURR AND H.FIRMNR=600
 LEFT JOIN LG_EXCHANGE_600 I WITH(NOLOCK) ON I.EDATE=A.DATE_ AND I.CRTYPE=20
 LEFT JOIN LG_600_01_INVOICE N1 WITH(NOLOCK) ON N1.LOGICALREF = A.SOURCEFREF
 LEFT JOIN LG_600_01_INVOICE N2 WITH(NOLOCK) ON N2.ACCFICHEREF = AA.LOGICALREF
 LEFT JOIN LG_600_CLCARD CL WITH(NOLOCK)  ON CL.LOGICALREF = N2.CLIENTREF
 LEFT JOIN (SELECT STL.INVOICEREF, SC.CANDEDUCT, STL.DEDUCTIONPART1,STL.DEDUCTIONPART2, STL.VAT,
            CASE
            WHEN (STL.IOCODE=1 OR STL.IOCODE=2 OR STL.TRCODE=4) OR (STL.IOCODE=0 AND STL.TRCODE IN (1,3)) THEN ROUND(STL.VATAMNT,2)
            WHEN (STL.IOCODE=3 OR STL.IOCODE=4 OR STL.TRCODE=9) OR (STL.IOCODE=0 AND STL.TRCODE IN (6,8)) THEN (-1)*ROUND(STL.VATAMNT,2)
            ELSE 0
            END AS VATAMOUNT,
            CASE
              WHEN (STL.IOCODE=1 OR STL.IOCODE=2 OR STL.TRCODE=4) OR (STL.IOCODE=0 AND STL.TRCODE IN (1,3)) THEN ROUND((STL.LINENET-(STL.DISTEXP-STL.DISTDISC)),2)
              WHEN (STL.IOCODE=3 OR STL.IOCODE=4 OR STL.TRCODE=9) OR (STL.IOCODE=0 AND STL.TRCODE IN (6,8)) THEN (-1)*ROUND((STL.LINENET-(STL.DISTEXP-STL.DISTDISC)),2)
              ELSE 0
              END AS GROSSTOTAL
            FROM LG_600_01_STLINE STL 
            JOIN LG_600_SRVCARD SC WITH(NOLOCK) ON SC.LOGICALREF=STL.STOCKREF
            WHERE STL.LINETYPE = 4
            AND SC.CANDEDUCT=1
            AND (STL.DEDUCTIONPART1 != 0 AND STL.DEDUCTIONPART2 !=0 )
            ) SSTL ON SSTL.INVOICEREF = N1.LOGICALREF 
INNER JOIN (
            SELECT EML.ACCFICHEREF, CREDIT AS VATAMOUNT, F.CODE AS VATCODE,
                    CASE 
                        WHEN F.CODE = '360.10.04.020' THEN '2/10'
                        WHEN F.CODE = '360.10.04.030' THEN '3/10'
                        WHEN F.CODE = '360.10.04.040' THEN '4/10'
                        WHEN F.CODE = '360.10.04.050' THEN '5/10'
                        WHEN F.CODE = '360.10.04.070' THEN '7/10'
                        WHEN F.CODE = '360.10.04.080' THEN '8/10'
                        WHEN F.CODE = '360.10.04.090' THEN '9/10'
                        WHEN F.CODE = '360.10.04.100' THEN '10/10'
                    END AS DEDUCTION            
            FROM  LG_600_01_EMFLINE EML WITH(NOLOCK)
             LEFT JOIN LG_600_01_EMFICHE AA WITH(NOLOCK) ON AA.LOGICALREF=EML.ACCFICHEREF
             LEFT JOIN LG_600_01_ACCDISTDETLN A1 WITH(NOLOCK) ON A1.PREVLINEREF=EML.LOGICALREF
             LEFT JOIN LG_600_EMUHACC F WITH(NOLOCK) ON F.LOGICALREF=EML.ACCOUNTREF
            WHERE AA.CANCELLED = 0 
            AND F.CODE LIKE '360.10.04%'
            AND AA.MODULENR=2
            ) SEML ON SEML.ACCFICHEREF = A.ACCFICHEREF
WHERE ISNULL(AA.CANCELLED,0) = 0 
AND ISNULL(A.CANCELLED,0)=0
AND (F.CODE LIKE '7%' OR F.CODE LIKE '253%' OR F.CODE LIKE '255%' OR F.CODE LIKE '260%')
AND MONTH(A.DATE_)= %s
AND YEAR(A.DATE_)= %s
AND AA.MODULENR=2
"""
            
            cursor.execute(query, (int(self.month), self.year))
            
            # Sonuçları Odoo'ya kaydet
            records = []
            for row in cursor:
                vals = {
                    'logo_id': row.get('logoID'),
                    'ay': row.get('ay'),
                    'yil': row.get('yil'),
                    'fis_no': row.get('fisNo'),
                    'proje': row.get('proje'),
                    'kebir_hesap_kodu': row.get('kebirHesapKodu'),
                    'kebir_hesap_adi': row.get('kebirHesapAdi'),
                    'hesap_kodu': row.get('hesapKodu'),
                    'hesap_adi': row.get('hesapAdi'),
                    'masraf_merkezi': row.get('masrafMerkezi'),
                    'kaynak_modul': row.get('kaynakModul'),
                    'aciklama': row.get('aciklama'),
                    'fis_aciklama': row.get('fisAciklama'),
                    'cari': row.get('cari'),
                    'cari_vergi_no': row.get('cariVergiNo'),
                    'cari_unvan': row.get('cariUnvan'),
                    'adi': row.get('adi'),
                    'soy_adi': row.get('soyAdi'),
                    'tckn': row.get('tckn'),
                    'tutar_yerel': row.get('tutarYerel') or 0.0,
                    'kdv_tutar': row.get('kdvTutar') or 0.0,
                    'tevkifat_oran': row.get('tevkifatOran'),
                    'tevkif_edilen_kdv_tutari': row.get('tevkifEdilenKdvTutari') or 0.0,
                }
                records.append(self.env['logo.kdv2.report'].create(vals))
            
            if records:
                # Rapor görünümünü aç
                return {
                    'name': _('KDV-2 Listesi - %s/%s') % (self.month, self.year),
                    'type': 'ir.actions.act_window',
                    'res_model': 'logo.kdv2.report',
                    'view_mode': 'list',
                    'domain': [],
                }
            else:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': _('Bilgi'),
                        'message': _('Seçilen dönem için kayıt bulunamadı.'),
                        'type': 'warning',
                        'sticky': False,
                    }
                }

        except Exception as e:
            _logger.error("KDV-2 rapor hatası: %s", str(e))
            raise UserError(_("Rapor oluşturma hatası: %s") % str(e))
        finally:
            # Bağlantıyı kapat
            if cursor:
                cursor.close()
            if conn:
                conn.close()


class LogoMuhtasarReport(models.TransientModel):
    _name = 'logo.muhtasar.report'
    _description = 'Logo Muhtasar Raporu'
    
    # Rapor sonuçları için alanlar
    odenecek_gelir_vergileri = fields.Char(string='Ödenecek Gelir Vergileri', readonly=True)
    vergi_turu = fields.Char(string='Vergi Türü', readonly=True)
    tarih = fields.Date(string='Tarih', readonly=True)
    ay = fields.Integer(string='Ay', readonly=True)
    yil = fields.Integer(string='Yıl', readonly=True)
    fis_no = fields.Char(string='Fiş No', readonly=True)
    islem = fields.Char(string='İşlem', readonly=True)
    is_yeri = fields.Char(string='İş Yeri', readonly=True)
    bolum = fields.Char(string='Bölüm', readonly=True)
    proje = fields.Char(string='Proje', readonly=True)
    kebir_hesabi_kodu = fields.Char(string='Kebir Hesabı Kodu', readonly=True)
    kebir_hesabi_adi = fields.Char(string='Kebir Hesabı Adı', readonly=True)
    hesap_kodu = fields.Char(string='Hesap Kodu', readonly=True)
    hesap_adi = fields.Char(string='Hesap Adı', readonly=True)
    masraf_merkezi = fields.Char(string='Masraf Merkezi', readonly=True)
    kaynak_modul = fields.Char(string='Kaynak Modül', readonly=True)
    tutar = fields.Float(string='Tutar', digits=(16, 2), readonly=True)
    tutar_yerel = fields.Float(string='Tutar (Yerel)', digits=(16, 2), readonly=True)
    aciklama = fields.Char(string='Açıklama', readonly=True)
    fis_aciklama = fields.Char(string='Fiş Açıklama', readonly=True)
    hareket_yonu = fields.Char(string='Hareket Yönü', readonly=True)
    iptal = fields.Char(string='İptal', readonly=True)
    belge_turu = fields.Char(string='Belge Türü', readonly=True)
    cari = fields.Char(string='Cari', readonly=True)
    cari_vergi_no = fields.Char(string='Cari Vergi No', readonly=True)
    cari_unvan1 = fields.Char(string='Cari Ünvan 1', readonly=True)
    cari_unvan2 = fields.Char(string='Cari Ünvan 2', readonly=True)
    adi = fields.Char(string='Adı', readonly=True)
    soyadi = fields.Char(string='Soyadı', readonly=True)
    fatura_belge_no = fields.Char(string='Fatura Belge No', readonly=True)
    fatura_no = fields.Char(string='Fatura No', readonly=True)
    adres1 = fields.Char(string='Adres', readonly=True)
    ulke = fields.Char(string='Ülke', readonly=True)
    vergi = fields.Char(string='Vergi', readonly=True)


class LogoMuhtasarWizard(models.TransientModel):
    _name = 'logo.muhtasar.wizard'
    _description = 'Logo Muhtasar Rapor Sihirbazı'
    
    month = fields.Selection([
        ('1', 'Ocak'),
        ('2', 'Şubat'),
        ('3', 'Mart'),
        ('4', 'Nisan'),
        ('5', 'Mayıs'),
        ('6', 'Haziran'),
        ('7', 'Temmuz'),
        ('8', 'Ağustos'),
        ('9', 'Eylül'),
        ('10', 'Ekim'),
        ('11', 'Kasım'),
        ('12', 'Aralık'),
    ], string='Ay', required=True, default=str(fields.Date.today().month))
    
    year = fields.Integer(string='Yıl', required=True, default=fields.Date.today().year)
    
    def _get_mssql_connection(self):
        """MSSQL bağlantısı oluştur"""
        if not pymssql:
            raise UserError(_("pymssql kütüphanesi yüklü değil. 'pip install pymssql' komutunu çalıştırın."))
        
        config_param = self.env['ir.config_parameter'].sudo()
        server = config_param.get_param('logo.mssql_server')
        port = int(config_param.get_param('logo.mssql_port', '1433'))
        database = config_param.get_param('logo.mssql_database')
        username = config_param.get_param('logo.mssql_username')
        password = config_param.get_param('logo.mssql_password')
        
        if not all([server, database, username, password]):
            raise UserError(_("Logo MSSQL bağlantı parametreleri eksik. Lütfen E-Fatura → Yapılandırma menüsünden ayarları tamamlayın."))
        
        try:
            connection = pymssql.connect(
                server=server,
                port=port,
                user=username,
                password=password,
                database=database,
                timeout=120,
                login_timeout=60,
                charset='UTF-8',
                appname='Odoo E-Fatura'
            )
            return connection
        except Exception as e:
            raise UserError(_("MSSQL bağlantı hatası: %s") % str(e))

    def action_generate_report(self):
        """Muhtasar raporunu oluştur"""
        conn = None
        cursor = None
        try:
            # Mevcut kayıtları temizle
            self.env['logo.muhtasar.report'].search([]).unlink()

            # MSSQL bağlantısı
            conn = self._get_mssql_connection()
            cursor = conn.cursor(as_dict=True)
            
            # SQL sorgusunu çalıştır
            query = """
SELECT 
CASE
    WHEN F.CODE = '360.10.01.001' THEN 'ÜCRET GELİR VERGİSİ'
    WHEN F.CODE = '360.10.01.002' THEN 'S.M MAKBUZU'
    WHEN F.CODE = '360.10.01.003' THEN 'KİRA GELİR VERGİSİ'
    WHEN F.CODE = '360.10.01.004' THEN 'GİDER PUSULASI GELİR VERGİSİ'
    WHEN F.CODE = '360.10.01.005' THEN 'YURT DIŞI HİZMERT ALIMI GELİR VERGİSİ'
    WHEN F.CODE LIKE '7%' THEN 'VERGİ' 
END as odenecekGelirVergileri,
CASE
    WHEN F.CODE = '360.10.01.001' THEN 'ÜCRET GELİR VERGİSİ'
    WHEN F.CODE = '360.10.01.002' THEN '022'
    WHEN F.CODE = '360.10.01.003' THEN '041'
    WHEN F.CODE = '360.10.01.004' THEN '156'
    WHEN F.CODE = '360.10.01.005' THEN '279'
    WHEN F.CODE LIKE '7%' THEN ''
END as vergiTuru,
A.DATE_ as tarih,
MONTH(A.DATE_) as ay,
YEAR(A.DATE_) as yil,
AA.FICHENO as fisNo,
CASE 
    WHEN A.TRCODE=1 THEN '1 Açılış'
    WHEN A.TRCODE=2 THEN '2 Tahsil'
    WHEN A.TRCODE=3 THEN '3 Tediye'
    WHEN A.TRCODE=4 THEN '4 Mahsup'
    WHEN A.TRCODE=5 THEN '5 Özel'
    WHEN A.TRCODE=6 THEN '6 Kur Farkı'
    WHEN A.TRCODE=7 THEN '7 Kapanış'
    ELSE '' 
END as islem,
CAST(C.NR AS CHAR(3))+' '+C.NAME as isYeri,
CAST(D.NR AS CHAR(3))+' '+D.NAME as bolum,
E.CODE+' '+E.NAME as proje,
F1.CODE as kebirHesabiKodu,
F1.DEFINITION_ as kebirHesabiAdi,
F.CODE as hesapKodu,
F.DEFINITION_ as hesapAdi,
G.CODE+' '+G.DEFINITION_ as masrafMerkezi,
CASE 
    WHEN AA.MODULENR=1 THEN '1 Malzeme'
    WHEN AA.MODULENR=2 THEN '2 Satınalma'
    WHEN AA.MODULENR=3 THEN '3 Satış'
    WHEN AA.MODULENR=4 THEN '4 Cari Hesap'
    WHEN AA.MODULENR=5 THEN '5 Çek Senet'
    WHEN AA.MODULENR=6 THEN '6 Banka'
    WHEN AA.MODULENR=7 THEN '7 Kasa'
    ELSE '' 
END as kaynakModul,
-CASE 
    WHEN A.TRCURR=0 AND (A.LOGICALREF NOT IN (SELECT DISTINCT PREVLINEREF FROM LG_600_01_ACCDISTDETLN) OR A1.DISTRATE=1) THEN ABS(A.DEBIT-A.CREDIT)-ABS(A.DEBIT-A.CREDIT)*2*A.SIGN
    WHEN A.TRCURR<>0 AND (A.LOGICALREF NOT IN (SELECT DISTINCT PREVLINEREF FROM LG_600_01_ACCDISTDETLN) OR A1.DISTRATE=1) THEN A.TRNET-A.TRNET*2*A.SIGN
    WHEN A.TRCURR=0 AND A1.DISTRATE<>100 THEN A1.CREDEBNET-A1.CREDEBNET*2*A.SIGN
    WHEN A.TRCURR<>0 AND A1.DISTRATE<>100 THEN A1.TRNET-A1.TRNET*2*A.SIGN
    ELSE 0 
END  as tutar,
CASE 
    WHEN A.LOGICALREF NOT IN (SELECT DISTINCT PREVLINEREF FROM LG_600_01_ACCDISTDETLN) OR A1.DISTRATE=1 THEN ABS(A.DEBIT-A.CREDIT)-ABS(A.DEBIT-A.CREDIT)*2*A.SIGN 
    ELSE A1.CREDEBNET-A1.CREDEBNET*2*A.SIGN 
END as tutarYerel,
A.LINEEXP as aciklama,
AA.GENEXP1 as fisAciklama,
CASE 
    WHEN A.SIGN=0 THEN '0 Borç' WHEN A.SIGN=1 THEN '1 Alacak' 
    ELSE '' 
END as hareketYonu,
CASE 
    WHEN A.CANCELLED=0 THEN 'Hayır' 
    ELSE 'Evet' 
END as iptal,
CASE AA.DOCTYPE 
    WHEN 0 THEN 'Normal' 
    WHEN 1 THEN 'Cost Of Sales' 
    WHEN 2 THEN 'Differences Of Cost Of Sales' 
    ELSE '' 
END as belgeTuru,
A.CLDEF as cari,
A.TAXNR as cariVergiNo,
CL.DEFINITION_ as cariUnvan1, 
CL.DEFINITION2 as cariUnvan2,
CL.NAME as adi, 
CL.SURNAME as soyadi,
N2.DOCODE as faturaBelgeNo, 
N2.FICHENO as faturaNo,
CL.ADDR1 as adres1,
CL.COUNTRY as ulke,
-1 * SEML.TOTAL as vergi
FROM  LG_600_01_EMFLINE A WITH(NOLOCK)
    LEFT JOIN LG_600_01_EMFICHE AA WITH(NOLOCK) ON AA.LOGICALREF=A.ACCFICHEREF
    LEFT JOIN LG_600_01_ACCDISTDETLN A1 WITH(NOLOCK) ON A1.PREVLINEREF=A.LOGICALREF
    LEFT JOIN L_CAPIDIV C WITH(NOLOCK) ON C.NR=A.BRANCH AND C.FIRMNR=600
    LEFT JOIN L_CAPIDEPT D WITH(NOLOCK) ON D.NR=A.DEPARTMENT AND D.FIRMNR=600
    LEFT JOIN LG_600_PROJECT E WITH(NOLOCK) ON E.LOGICALREF=A1.PROJECTREF
    LEFT JOIN LG_600_EMUHACC F WITH(NOLOCK) ON F.LOGICALREF=A.ACCOUNTREF
    LEFT JOIN LG_600_EMUHACC F1 WITH(NOLOCK) ON F1.CODE=left(F.CODE,3)
    LEFT JOIN LG_600_EMCENTER G WITH(NOLOCK) ON G.LOGICALREF=A.CENTERREF
    LEFT JOIN L_CURRENCYLIST H WITH(NOLOCK) ON H.CURTYPE=A.TRCURR AND H.FIRMNR=600
    LEFT JOIN LG_EXCHANGE_600 I WITH(NOLOCK) ON I.EDATE=A.DATE_ AND I.CRTYPE=20
    LEFT JOIN LG_600_01_INVOICE N1 WITH(NOLOCK) ON N1.LOGICALREF = A.SOURCEFREF
    LEFT JOIN LG_600_01_INVOICE N2 WITH(NOLOCK) ON N2.ACCFICHEREF = AA.LOGICALREF
    LEFT JOIN LG_600_CLCARD CL WITH(NOLOCK)  ON CL.LOGICALREF = N2.CLIENTREF
    INNER JOIN (
                SELECT
                EML.ACCFICHEREF, 
                CASE 
                    WHEN EML.LOGICALREF NOT IN (SELECT DISTINCT PREVLINEREF FROM LG_600_01_ACCDISTDETLN) OR A1.DISTRATE=1 THEN ABS(EML.DEBIT-EML.CREDIT)-ABS(EML.DEBIT-EML.CREDIT)*2*EML.SIGN 
                    ELSE A1.CREDEBNET-A1.CREDEBNET*2*EML.SIGN 
                END AS TOTAL
                FROM  LG_600_01_EMFLINE EML WITH(NOLOCK)
                    LEFT JOIN LG_600_01_EMFICHE AA WITH(NOLOCK) ON AA.LOGICALREF=EML.ACCFICHEREF
                    LEFT JOIN LG_600_01_ACCDISTDETLN A1 WITH(NOLOCK) ON A1.PREVLINEREF=EML.LOGICALREF
                    LEFT JOIN LG_600_EMUHACC F WITH(NOLOCK) ON F.LOGICALREF=EML.ACCOUNTREF
                WHERE AA.CANCELLED = 0
                    AND (F.CODE LIKE '360.10.01%')
                    AND MONTH(EML.DATE_)= %s                                      
                    AND YEAR(EML.DATE_)= %s                     
                    and AA.MODULENR=2
                ) SEML ON  SEML.ACCFICHEREF = A.ACCFICHEREF
WHERE AA.CANCELLED = 0
    AND (F.CODE LIKE '7%')
    AND MONTH(A.DATE_)= %s                 
    AND YEAR(A.DATE_)= %s                   
    and AA.MODULENR=2
UNION ALL
SELECT 
CASE
    WHEN F.CODE = '360.10.01.001' THEN 'ÜCRET GELİR VERGİSİ'
    WHEN F.CODE = '360.10.01.002' THEN 'S.M MAKBUZU'
    WHEN F.CODE = '360.10.01.003' THEN 'KİRA GELİR VERGİSİ'
    WHEN F.CODE = '360.10.01.004' THEN 'GİDER PUSULASI GELİR VERGİSİ'
    WHEN F.CODE = '360.10.01.005' THEN 'YURT DIŞI HİZMERT ALIMI GELİR VERGİSİ'
    WHEN F.CODE LIKE '7%' THEN 'VERGİ'
END as odenecekGelirVergileri,
CASE
    WHEN F.CODE = '360.10.01.001' THEN 'ÜCRET GELİR VERGİSİ'
    WHEN F.CODE = '360.10.01.002' THEN '022'
    WHEN F.CODE = '360.10.01.003' THEN '041'
    WHEN F.CODE = '360.10.01.004' THEN '156'
    WHEN F.CODE = '360.10.01.005' THEN '279'
    WHEN F.CODE LIKE '7%' THEN ''
END as vergiTuru,
A.DATE_ as tarih,
MONTH(A.DATE_) as ay,
YEAR(A.DATE_) as yil,
AA.FICHENO as fisNo,
CASE 
    WHEN A.TRCODE=1 THEN '1 Açılış'
    WHEN A.TRCODE=2 THEN '2 Tahsil'
    WHEN A.TRCODE=3 THEN '3 Tediye'
    WHEN A.TRCODE=4 THEN '4 Mahsup'
    WHEN A.TRCODE=5 THEN '5 Özel'
    WHEN A.TRCODE=6 THEN '6 Kur Farkı'
    WHEN A.TRCODE=7 THEN '7 Kapanış'
    ELSE '' 
END as islem,
CAST(C.NR AS CHAR(3))+' '+C.NAME as isYeri,
CAST(D.NR AS CHAR(3))+' '+D.NAME as bolum,
E.CODE+' '+E.NAME as proje,
F1.CODE as kebirHesabiKodu,
F1.DEFINITION_ as kebirHesabiAdi,
F.CODE as hesapKodu,
F.DEFINITION_ as hesapAdi,
G.CODE+' '+G.DEFINITION_ as masrafMerkezi,
CASE 
WHEN AA.MODULENR=1 THEN '1 Malzeme'
    WHEN AA.MODULENR=2 THEN '2 Satınalma'
    WHEN AA.MODULENR=3 THEN '3 Satış'
    WHEN AA.MODULENR=4 THEN '4 Cari Hesap'
    WHEN AA.MODULENR=5 THEN '5 Çek Senet'
    WHEN AA.MODULENR=6 THEN '6 Banka'
    WHEN AA.MODULENR=7 THEN '7 Kasa'
    ELSE '' 
END as kaynakModul,
-CASE 
    WHEN A.TRCURR=0 AND (A.LOGICALREF NOT IN (SELECT DISTINCT PREVLINEREF FROM LG_600_01_ACCDISTDETLN) OR A1.DISTRATE=1) THEN ABS(A.DEBIT-A.CREDIT)-ABS(A.DEBIT-A.CREDIT)*2*A.SIGN
    WHEN A.TRCURR<>0 AND (A.LOGICALREF NOT IN (SELECT DISTINCT PREVLINEREF FROM LG_600_01_ACCDISTDETLN) OR A1.DISTRATE=1) THEN A.TRNET-A.TRNET*2*A.SIGN
    WHEN A.TRCURR=0 AND A1.DISTRATE<>100 THEN A1.CREDEBNET-A1.CREDEBNET*2*A.SIGN
    WHEN A.TRCURR<>0 AND A1.DISTRATE<>100 THEN A1.TRNET-A1.TRNET*2*A.SIGN
    ELSE 0 
END as tutar,
CASE 
    WHEN A.LOGICALREF NOT IN (SELECT DISTINCT PREVLINEREF FROM LG_600_01_ACCDISTDETLN) OR A1.DISTRATE=1 THEN ABS(A.DEBIT-A.CREDIT)-ABS(A.DEBIT-A.CREDIT)*2*A.SIGN 
    ELSE A1.CREDEBNET-A1.CREDEBNET*2*A.SIGN 
END as tutarYerel,
A.LINEEXP as aciklama,
AA.GENEXP1 as fisAciklama,
CASE 
    WHEN A.SIGN=0 THEN '0 Borç' 
    WHEN A.SIGN=1 THEN '1 Alacak' 
    ELSE '' 
END as hareketYonu,
CASE 
    WHEN A.CANCELLED=0 THEN 'Hayır' 
    ELSE 'Evet' 
END  as iptal,
CASE AA.DOCTYPE 
    WHEN 0 THEN 'Normal' 
    WHEN 1 THEN 'Cost Of Sales' 
    WHEN 2 THEN 'Differences Of Cost Of Sales' 
    ELSE '' 
END as belgeTuru,
A.CLDEF as cari, 
A.TAXNR as cariVergiNo,
CL.DEFINITION_ as cariUnvan1, 
CL.DEFINITION2 as cariUnvan2,
CL.NAME as adi, 
CL.SURNAME as soyadi,
N2.DOCODE as faturaBelgeNo, 
N2.FICHENO as faturaNo,
CL.ADDR1 as adres1,
CL.COUNTRY as ulke,
-1 * SEML.TOTAL as vergi
FROM  LG_600_01_EMFLINE A WITH(NOLOCK)
    LEFT JOIN LG_600_01_EMFICHE AA WITH(NOLOCK) ON AA.LOGICALREF=A.ACCFICHEREF
    LEFT JOIN LG_600_01_ACCDISTDETLN A1 WITH(NOLOCK) ON A1.PREVLINEREF=A.LOGICALREF
    LEFT JOIN L_CAPIDIV C WITH(NOLOCK) ON C.NR=A.BRANCH AND C.FIRMNR=600
    LEFT JOIN L_CAPIDEPT D WITH(NOLOCK) ON D.NR=A.DEPARTMENT AND D.FIRMNR=600
    LEFT JOIN LG_600_PROJECT E WITH(NOLOCK) ON E.LOGICALREF=A1.PROJECTREF
    LEFT JOIN LG_600_EMUHACC F WITH(NOLOCK) ON F.LOGICALREF=A.ACCOUNTREF
    LEFT JOIN LG_600_EMUHACC F1 WITH(NOLOCK) ON F1.CODE=left(F.CODE,3)
    LEFT JOIN LG_600_EMCENTER G WITH(NOLOCK) ON G.LOGICALREF=A.CENTERREF
    LEFT JOIN L_CURRENCYLIST H WITH(NOLOCK) ON H.CURTYPE=A.TRCURR AND H.FIRMNR=600
    LEFT JOIN LG_EXCHANGE_600 I WITH(NOLOCK) ON I.EDATE=A.DATE_ AND I.CRTYPE=20
    LEFT JOIN LG_600_01_INVOICE N1 WITH(NOLOCK) ON N1.LOGICALREF = A.SOURCEFREF
    LEFT JOIN LG_600_01_INVOICE N2 WITH(NOLOCK) ON N2.ACCFICHEREF = AA.LOGICALREF
    INNER JOIN LG_600_CLCARD CL WITH(NOLOCK)  ON CL.LOGICALREF = N2.CLIENTREF
    INNER JOIN (
                SELECT
                EML.ACCFICHEREF, 
                CASE
                    WHEN EML.LOGICALREF NOT IN (SELECT DISTINCT PREVLINEREF FROM LG_600_01_ACCDISTDETLN) OR A1.DISTRATE=1 THEN ABS(EML.DEBIT-EML.CREDIT)-ABS(EML.DEBIT-EML.CREDIT)*2*EML.SIGN 
                    ELSE A1.CREDEBNET-A1.CREDEBNET*2*EML.SIGN 
                END AS TOTAL
                FROM  LG_600_01_EMFLINE EML WITH(NOLOCK)
                    LEFT JOIN LG_600_01_EMFICHE AA WITH(NOLOCK) ON AA.LOGICALREF=EML.ACCFICHEREF
                    LEFT JOIN LG_600_01_ACCDISTDETLN A1 WITH(NOLOCK) ON A1.PREVLINEREF=EML.LOGICALREF
                    LEFT JOIN LG_600_EMUHACC F WITH(NOLOCK) ON F.LOGICALREF=EML.ACCOUNTREF
                WHERE AA.CANCELLED = 0
                    AND (F.CODE LIKE '360.10.01%')
                    AND MONTH(EML.DATE_)= %s   
                    AND YEAR(EML.DATE_)= %s          
                    AND AA.MODULENR=6
                ) SEML ON  SEML.ACCFICHEREF = A.ACCFICHEREF
WHERE AA.CANCELLED = 0
    AND (F.CODE LIKE '740.YÜ[PM]%' OR F.CODE LIKE '770.10.08.001')
    AND MONTH(A.DATE_)= %s              
    AND YEAR(A.DATE_)= %s                  
    AND AA.MODULENR=6
"""
            
            cursor.execute(query, (int(self.month), self.year, int(self.month), self.year, int(self.month), self.year, int(self.month), self.year, int(self.month), self.year, int(self.month), self.year, int(self.month), self.year, int(self.month), self.year))
            
            # Sonuçları Odoo'ya kaydet
            records = []
            for row in cursor:
                vals = {
                    'odenecek_gelir_vergileri': row.get('odenecekGelirVergileri'),
                    'vergi_turu': row.get('vergiTuru'),
                    'tarih': row.get('tarih'),
                    'ay': row.get('ay'),
                    'yil': row.get('yil'),
                    'fis_no': row.get('fisNo'),
                    'islem': row.get('islem'),
                    'is_yeri': row.get('isYeri'),
                    'bolum': row.get('bolum'),
                    'proje': row.get('proje'),
                    'kebir_hesabi_kodu': row.get('kebirHesabiKodu'),
                    'kebir_hesabi_adi': row.get('kebirHesabiAdi'),
                    'hesap_kodu': row.get('hesapKodu'),
                    'hesap_adi': row.get('hesapAdi'),
                    'masraf_merkezi': row.get('masrafMerkezi'),
                    'kaynak_modul': row.get('kaynakModul'),
                    'tutar': row.get('tutar') or 0.0,
                    'tutar_yerel': row.get('tutarYerel') or 0.0,
                    'aciklama': row.get('aciklama'),
                    'fis_aciklama': row.get('fisAciklama'),
                    'hareket_yonu': row.get('hareketYonu'),
                    'iptal': row.get('iptal'),
                    'belge_turu': row.get('belgeTuru'),
                    'cari': row.get('cari'),
                    'cari_vergi_no': row.get('cariVergiNo'),
                    'cari_unvan1': row.get('cariUnvan1'),
                    'cari_unvan2': row.get('cariUnvan2'),
                    'adi': row.get('adi'),
                    'soyadi': row.get('soyadi'),
                    'fatura_belge_no': row.get('faturaBelgeNo'),
                    'fatura_no': row.get('faturaNo'),
                    'adres1': row.get('adres1'),
                    'ulke': row.get('ulke'),
                }
                records.append(self.env['logo.muhtasar.report'].create(vals))

            if records:
                # Rapor görünümünü aç
                return {
                    'name': _('Muhtasar Listesi - %s/%s') % (self.month, self.year),
                    'type': 'ir.actions.act_window',
                    'res_model': 'logo.muhtasar.report',
                    'view_mode': 'list',
                    'domain': [],
                }
            else:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': _('Bilgi'),
                        'message': _('Seçilen dönem için kayıt bulunamadı.'),
                        'type': 'warning',
                        'sticky': False,
                    }
                }

        except Exception as e:
            _logger.error("Muhtasar rapor hatası: %s", str(e))
            raise UserError(_("Rapor oluşturma hatası: %s") % str(e))
        finally:
            # Bağlantıyı kapat
            if cursor:
                cursor.close()
            if conn:
                conn.close()


class EarsivExcelImportWizard(models.TransientModel):
    _name = 'earsiv.excel.import.wizard'
    _description = 'E-Arşiv Gelen Excel Import Wizard'

    # Dosya yükleme alanı
    file_data = fields.Binary('Excel Dosyası', required=True)
    file_name = fields.Char('Dosya Adı')

    # Import özeti
    import_summary = fields.Html('Import Özeti', readonly=True)

    # Logo sync otomatik
    auto_logo_sync = fields.Boolean('Logo Senkronizasyonu Yap', default=True,
                                   help="Import sonrası otomatik Logo senkronizasyonu yapılsın mı?")

    def validate_excel_format(self, worksheet):
        """Excel formatını doğrula"""
        required_headers = [
            'Sıra',
            'Ünvanı/Adı Soyadı',
            'Vergi Kimlik/T.C. Kimlik Numarası',
            'Fatura No',
            'Düzenleme Tarihi',
            'Toplam Tutar',
            'Ödenecek Tutar',
            'Vergiler Toplamı',
            'Para Birimi',
            'Tesisat Numarası',
            'Gönderim Şekli',
            'İptal İtiraz Durum',
            'İptal İtiraz Tarihi'
        ]

        # Header satırını oku (row 2)
        headers = []
        for col in range(1, 14):
            cell_value = worksheet.cell(row=2, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())

        # Eksik başlıkları kontrol et
        missing_headers = []
        for required in required_headers:
            if required not in headers:
                missing_headers.append(required)

        if missing_headers:
            raise UserError(_(
                "❌ HATA: Excel dosyası beklenen formatta değil!\n\n"
                "📋 Eksik sütunlar:\n%s\n\n"
                "✅ Beklenen format:\n"
                "1. Sıra\n"
                "2. Ünvanı/Adı Soyadı\n"
                "3. Vergi Kimlik/T.C. Kimlik Numarası\n"
                "4. Fatura No\n"
                "5. Düzenleme Tarihi\n"
                "6. Toplam Tutar\n"
                "7. Ödenecek Tutar\n"
                "8. Vergiler Toplamı\n"
                "9. Para Birimi\n"
                "10. Tesisat Numarası\n"
                "11. Gönderim Şekli\n"
                "12. İptal İtiraz Durum\n"
                "13. İptal İtiraz Tarihi\n\n"
                "Lütfen doğru formatta bir Excel dosyası yükleyin!"
            ) % '\n'.join(['• ' + h for h in missing_headers]))

        return True

    def _parse_excel_date(self, date_value):
        """Excel tarih formatını parse et"""
        if not date_value:
            return None

        if isinstance(date_value, datetime):
            return date_value

        # String ise parse et
        date_str = str(date_value).strip()

        # Farklı formatları dene
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%d/%m/%Y',
            '%d-%m-%Y',
            '%d.%m.%Y',
            '%Y-%m-%d',
        ]

        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except:
                continue

        _logger.warning("E-Arşiv Excel Import: Tarih parse edilemedi: %s", date_str)
        return None

    def _create_or_update_invoice(self, row_data):
        """e.invoice kaydı oluştur veya güncelle"""
        Invoice = self.env['e.invoice']

        # Mevcut kaydı kontrol et
        existing = Invoice.search([
            ('invoice_id', '=', row_data['Fatura No']),
            ('kaynak', '=', 'e-arsiv'),
            ('direction', '=', 'IN')
        ], limit=1)

        # Değerleri hazırla
        vals = {
            'invoice_id': row_data['Fatura No'],
            'uuid': f"{row_data['Fatura No']}_EXCEL_{datetime.now().strftime('%Y%m%d%H%M%S')}",
            'direction': 'IN',
            'kaynak': 'e-arsiv',
            'sender': str(row_data.get('Vergi Kimlik/T.C. Kimlik Numarası', '')),
            'receiver': '4510016851',  # Bizim VKN
            'supplier': row_data.get('Ünvanı/Adı Soyadı', ''),
            'customer': 'GÜVEN HASTANESİ A.Ş',
            'issue_date': self._parse_excel_date(row_data.get('Düzenleme Tarihi')),
            'tax_exclusive_total_amount': float(row_data.get('Toplam Tutar', 0) or 0),
            'payable_amount': float(row_data.get('Ödenecek Tutar', 0) or 0),
            'currency_code': row_data.get('Para Birimi', 'TRY'),
            'sending_type': row_data.get('Gönderim Şekli', ''),
            'profile_id': 'EARSIVFATURA',
            'status_code': '150' if 'İptal' in str(row_data.get('İptal İtiraz Durum', '')) else '130',
            'status_description': row_data.get('İptal İtiraz Durum', 'Raporlandı'),
            'notes': f"Excel Import - {datetime.now()}\nİptal Tarihi: {row_data.get('İptal İtiraz Tarihi', '')}",
        }

        # Vergiler toplamını hesapla
        if row_data.get('Vergiler Toplamı'):
            try:
                vergiler = float(row_data.get('Vergiler Toplamı', 0) or 0)
                vals['tax_inclusive_total_amount'] = vals['tax_exclusive_total_amount'] + vergiler
            except:
                vals['tax_inclusive_total_amount'] = vals['tax_exclusive_total_amount']

        if existing:
            # Kilitli kayıtları güncelleme (v1.0.5)
            if existing.is_locked:
                _logger.info(f"Excel Import: Kilitli kayıt atlandı - {row_data['Fatura No']}")
                return existing.with_context(is_new=False, skipped=True)

            # UUID'yi güncelleme durumunda değiştirme
            vals.pop('uuid', None)
            existing.write(vals)
            return existing.with_context(is_new=False)
        else:
            new_invoice = Invoice.create(vals)
            return new_invoice.with_context(is_new=True)

    def _run_logo_sync(self, invoice_ids):
        """Logo senkronizasyonu çalıştır"""
        try:
            logo_wizard = self.env['logo.sync.wizard'].create({
                'sync_mode': 'selected',
                'date_filter': False,
                'direction_filter': 'IN',
                'test_mode': False
            })
            logo_wizard.with_context(active_ids=invoice_ids).action_sync_logo()
            _logger.info("E-Arşiv Excel Import: Logo sync başarıyla tamamlandı")
        except Exception as e:
            _logger.warning("E-Arşiv Excel Import: Logo sync hatası: %s", str(e))

    def _prepare_summary(self, created, updated, errors):
        """Import özeti hazırla"""
        from markupsafe import Markup

        if errors > 0:
            alert_class = "alert-warning"
        else:
            alert_class = "alert-success"

        summary = f"""
        <div class="alert {alert_class}">
            <h4>✅ Import Tamamlandı!</h4>
            <ul>
                <li><strong>Yeni Kayıt:</strong> {created}</li>
                <li><strong>Güncellenen:</strong> {updated}</li>
                <li><strong>Hata:</strong> {errors}</li>
                <li><strong>Toplam İşlenen:</strong> {created + updated}</li>
            </ul>
        </div>
        """
        return Markup(summary)

    def action_import(self):
        """Excel dosyasını import et"""
        import base64
        import io

        try:
            import openpyxl
        except ImportError:
            raise UserError(_("openpyxl kütüphanesi yüklü değil. Sistem yöneticisine başvurun."))

        if not self.file_data:
            raise UserError(_("Lütfen bir Excel dosyası seçin!"))

        # Dosyayı decode et
        file_content = base64.b64decode(self.file_data)
        file_stream = io.BytesIO(file_content)

        try:
            # Excel dosyasını aç
            workbook = openpyxl.load_workbook(file_stream, read_only=True)

            # İlk sheet'i al
            worksheet = workbook.active

            # FORMAT DOĞRULAMA
            self.validate_excel_format(worksheet)

            # Veri işleme
            created_count = 0
            updated_count = 0
            error_count = 0
            invoice_ids = []

            # Satırları işle (row 3'ten başla - data rows)
            for row_num in range(3, worksheet.max_row + 1):
                try:
                    row_data = {}
                    row_data['Sıra'] = worksheet.cell(row=row_num, column=1).value
                    row_data['Ünvanı/Adı Soyadı'] = worksheet.cell(row=row_num, column=2).value
                    row_data['Vergi Kimlik/T.C. Kimlik Numarası'] = worksheet.cell(row=row_num, column=3).value
                    row_data['Fatura No'] = worksheet.cell(row=row_num, column=4).value
                    row_data['Düzenleme Tarihi'] = worksheet.cell(row=row_num, column=5).value
                    row_data['Toplam Tutar'] = worksheet.cell(row=row_num, column=6).value
                    row_data['Ödenecek Tutar'] = worksheet.cell(row=row_num, column=7).value
                    row_data['Vergiler Toplamı'] = worksheet.cell(row=row_num, column=8).value
                    row_data['Para Birimi'] = worksheet.cell(row=row_num, column=9).value
                    row_data['Tesisat Numarası'] = worksheet.cell(row=row_num, column=10).value
                    row_data['Gönderim Şekli'] = worksheet.cell(row=row_num, column=11).value
                    row_data['İptal İtiraz Durum'] = worksheet.cell(row=row_num, column=12).value
                    row_data['İptal İtiraz Tarihi'] = worksheet.cell(row=row_num, column=13).value

                    # Boş satırları atla
                    if not row_data['Fatura No']:
                        continue

                    # e.invoice kaydı oluştur/güncelle
                    invoice = self._create_or_update_invoice(row_data)
                    if invoice:
                        invoice_ids.append(invoice.id)
                        if invoice._context.get('is_new'):
                            created_count += 1
                        else:
                            updated_count += 1

                except Exception as e:
                    error_count += 1
                    _logger.error("E-Arşiv Excel Import - Satır %s import hatası: %s", row_num, str(e))

            workbook.close()

            # Logo sync
            if self.auto_logo_sync and invoice_ids:
                self._run_logo_sync(invoice_ids)

            # Özet mesajı
            self.import_summary = self._prepare_summary(created_count, updated_count, error_count)

            # Aynı wizard'ı göster (özet ile birlikte)
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'earsiv.excel.import.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }

        except openpyxl.utils.exceptions.InvalidFileException:
            raise UserError(_(
                "❌ HATA: Geçersiz Excel dosyası!\n\n"
                "Yüklediğiniz dosya geçerli bir Excel dosyası değil.\n"
                "Lütfen .xlsx uzantılı bir Excel dosyası yükleyin."
            ))
        except Exception as e:
            if "beklenen formatta değil" in str(e):
                raise  # Format hatası mesajını aynen göster
            else:
                raise UserError(_("Import hatası: %s") % str(e))